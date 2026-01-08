"""
Daily Planning Optimizer - Progressive Fill
Fills each day to 100% hours before moving to the next day.
Balances lines proportionally within each day.
"""
import csv
import openpyxl
import os
from datetime import datetime
from typing import List, Dict, Tuple
import json


class DailyPlanOptimizerProgressive:
    def __init__(self, template_path: str = None):
        """
        Initialize optimizer with Excel template file.
        
        Args:
            template_path: Path to Excel template file (default: 'Daily Planning Template.xlsm')
        """
        self.template_path = template_path or 'Daily Planning Template.xlsm'
        self.limits = {}
        self.orders = []
        self.brand_limits = {}
        
    def load_data(self):
        """Load orders and limits directly from Excel template."""
        self._load_limits()
        self._load_orders()
        
    def _load_limits(self):
        """Load brand-specific limits directly from Excel template."""
        if not os.path.exists(self.template_path):
            raise FileNotFoundError(f"Template file not found: {self.template_path}")
        
        wb = openpyxl.load_workbook(self.template_path, data_only=True)
        main_sheet = wb['Main']
        
        # Extract limits (row 2)
        headers = [cell.value for cell in main_sheet[1]]
        limits_row = [cell.value for cell in main_sheet[2]]
        
        limits = {}
        try:
            qty_idx = headers.index('Qty')
            picks_idx = headers.index('Picks')
            hours_idx = headers.index('Hours')
            limits = {
                'Qty': self._parse_float(limits_row[qty_idx]),
                'Picks': self._parse_float(limits_row[picks_idx]),
                'Hours': self._parse_float(limits_row[hours_idx])
            }
        except ValueError:
            print("Warning: Could not extract basic limits")
        
        # Extract brand-specific limits (BVI and Malosa)
        bvi_limits = {}
        malosa_limits = {}
        
        try:
            bvi_limits = {
                'Qty': limits.get('Qty', 0),
                'Picks': limits.get('Picks', 0),
                'Hours': limits.get('Hours', 0),
            }
            
            if 'Low Picks' in headers:
                try:
                    bvi_limits['Low Picks'] = self._parse_float(limits_row[headers.index('Low Picks')])
                except (ValueError, IndexError):
                    pass
            if 'Big Picks' in headers:
                try:
                    bvi_limits['Big Picks'] = self._parse_float(limits_row[headers.index('Big Picks')])
                except (ValueError, IndexError):
                    pass
            if 'Large Orders' in headers:
                try:
                    bvi_limits['Large Orders'] = self._parse_float(limits_row[headers.index('Large Orders')])
                except (ValueError, IndexError):
                    pass
            if 'Offline Jobs' in headers:
                try:
                    bvi_limits['Offline Jobs'] = self._parse_float(limits_row[headers.index('Offline Jobs')])
                except (ValueError, IndexError):
                    pass
            
            if 'Malosa' in headers:
                malosa_start = headers.index('Malosa')
                for i in range(malosa_start, len(headers)):
                    if headers[i] == 'Qty' and i + 2 < len(limits_row):
                        malosa_limits = {
                            'Qty': self._parse_float(limits_row[i]),
                            'Picks': self._parse_float(limits_row[i+1]),
                            'Hours': self._parse_float(limits_row[i+2]),
                        }
                        break
        except Exception as e:
            print(f"Warning: Could not extract brand-specific limits: {e}")
        
        if not bvi_limits.get('Qty'):
            bvi_limits = {'Qty': 10544, 'Picks': 750, 'Hours': 390}
        if not malosa_limits.get('Qty'):
            malosa_limits = {'Qty': 3335, 'Picks': 130, 'Hours': 90}
        
        self.brand_limits = {
            'BVI': bvi_limits,
            'Malosa': malosa_limits
        }
        
        print(f"Loaded brand limits: {self.brand_limits}")
        self.limits = self.brand_limits.get('BVI', {})
    
    def _load_orders(self):
        """Load orders directly from Excel template."""
        if not os.path.exists(self.template_path):
            raise FileNotFoundError(f"Template file not found: {self.template_path}")
        
        wb = openpyxl.load_workbook(self.template_path, data_only=True)
        main_sheet = wb['Main']
        
        order_headers = [cell.value for cell in main_sheet[11]]
        
        for row_idx in range(12, main_sheet.max_row + 1):
            row = [cell.value for cell in main_sheet[row_idx]]
            if not row[0] or row[0] == 'Order No':
                continue
            
            try:
                row_dict = {}
                for idx, header in enumerate(order_headers):
                    if idx < len(row) and header:
                        value = row[idx]
                        row_dict[header] = value
                
                suggested_line = str(row_dict.get('Suggested Line', '')).strip()
                if suggested_line in ['C3/4', 'C3&4']:
                    suggested_line = 'C3/4'
                
                qty = self._parse_float(row_dict.get('Lot Size', 0))
                picks = self._parse_float(row_dict.get('Picks', 0))
                hours = self._parse_float(row_dict.get('Hours', 0))
                
                start_date = row_dict.get('Start Date')
                if isinstance(start_date, datetime):
                    parsed_date = start_date
                elif start_date:
                    parsed_date = self._parse_date(str(start_date))
                else:
                    parsed_date = None
                
                # Load efficiency metrics for difficulty classification
                qty_hr = self._parse_float(row_dict.get('Qty/Hr', 0))
                picks_hr = self._parse_float(row_dict.get('Picks/Hr', 0))
                picks_qty = self._parse_float(row_dict.get('Picks/Qty', 0))
                
                # Calculate if not present in data
                if qty_hr == 0 and hours > 0:
                    qty_hr = qty / hours
                if picks_hr == 0 and hours > 0:
                    picks_hr = picks / hours
                if picks_qty == 0 and qty > 0:
                    picks_qty = picks / qty
                
                order = {
                    'Order No': str(row_dict.get('Order No', '')).strip(),
                    'Part No': str(row_dict.get('Part No', '')).strip(),
                    'Brand': str(row_dict.get('Brand', '')).strip(),
                    'Start Date': parsed_date,
                    'Lot Size': qty,
                    'Picks': picks,
                    'Hours': hours,
                    'Country': str(row_dict.get('Country', '')).strip(),
                    'Wrap Type': str(row_dict.get('Wrap Type', '')).strip(),
                    'CPU': self._parse_float(row_dict.get('CPU', 0)),
                    'Suggested Line': suggested_line,
                    'Qty/Hr': qty_hr,
                    'Picks/Hr': picks_hr,
                    'Picks/Qty': picks_qty,
                }
                
                if order['Order No'] and order['Part No'] and order['Lot Size'] > 0:
                    self.orders.append(order)
            except Exception as e:
                order_no = row[0] if row else 'unknown'
                print(f"Error processing order {order_no}: {e}")
                continue
        
        print(f"Loaded {len(self.orders)} orders")
    
    def _parse_date(self, date_str):
        """Parse date string to datetime object."""
        if not date_str or date_str.strip() == '':
            return None
        try:
            for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%Y-%m-%d %H:%M:%S']:
                try:
                    return datetime.strptime(date_str.strip(), fmt)
                except ValueError:
                    continue
            return None
        except:
            return None
    
    def _parse_float(self, value):
        """Parse float value, handling empty strings and None."""
        if value is None or value == '' or value == '-':
            return 0.0
        try:
            return float(str(value).replace(',', ''))
        except:
            return 0.0
    
    def _get_line_category(self, line: str) -> str:
        """Normalize line name to category (C1, C2, C3/4, or Other)."""
        if not line:
            return 'Other'
        line_upper = line.upper().strip()
        if line_upper in ['C1']:
            return 'C1'
        elif line_upper in ['C2']:
            return 'C2'
        elif line_upper in ['C3/4', 'C3&4', 'C3/4']:
            return 'C3/4'
        else:
            return 'Other'
    
    def _calculate_difficulty_score(self, order: Dict) -> float:
        """
        Calculate a normalized difficulty score for an order.
        
        Higher score = MORE DIFFICULT (slower, more complex)
        Lower score = EASIER (faster, simpler)
        
        Based on:
        - Qty/Hr: Higher is easier (more productive) -> inverted for score
        - Picks/Qty: Higher is harder (more complex) -> direct for score
        
        Returns a score normalized roughly around 0 (easy negative, hard positive)
        """
        qty_hr = order.get('Qty/Hr', 0)
        picks_qty = order.get('Picks/Qty', 0)
        
        # Normalize Qty/Hr (typical range 5-50, higher = easier)
        # Invert so higher difficulty score = harder
        if qty_hr > 0:
            qty_hr_score = 1.0 / qty_hr * 20  # Scale factor
        else:
            qty_hr_score = 2.0  # Default to hard if no data
        
        # Picks/Qty (typical range 0.1-1.0, higher = harder)
        picks_qty_score = picks_qty * 2  # Scale factor
        
        # Combined score (higher = harder)
        difficulty = qty_hr_score + picks_qty_score
        
        return difficulty
    
    def _classify_difficulty(self, score: float, thresholds: Dict) -> str:
        """Classify difficulty score into Easy/Medium/Hard."""
        if score <= thresholds['easy']:
            return 'Easy'
        elif score >= thresholds['hard']:
            return 'Hard'
        else:
            return 'Medium'
    
    def _calculate_difficulty_thresholds(self, orders_with_metrics: List[Dict]) -> Dict:
        """Calculate difficulty thresholds based on the distribution of orders."""
        scores = [item['difficulty_score'] for item in orders_with_metrics]
        scores.sort()
        
        n = len(scores)
        if n < 3:
            return {'easy': 0.5, 'hard': 1.5}
        
        # Use percentiles: bottom 33% = Easy, top 33% = Hard
        easy_threshold = scores[int(n * 0.33)]
        hard_threshold = scores[int(n * 0.67)]
        
        return {'easy': easy_threshold, 'hard': hard_threshold}
    
    def generate_multi_day_plans(self, num_days: int, brand: str = None) -> List[Dict]:
        """
        Generate multi-day plans with balanced hours AND order counts.
        
        MULTI-ROUND LEVELING APPROACH:
        1. Round-robin distribute orders across days for even order counts
        2. Perform swap-based rebalancing to level hours while keeping order counts similar
        3. Multiple passes to ensure both metrics are balanced
        4. Last day (remainder) is still allowed but should be reasonably balanced
        
        Args:
            num_days: Maximum number of days to plan
            brand: Brand to plan (BVI, Malosa, etc.)
        
        Returns:
            List of day plans with 'day', 'orders', 'totals', 'utilization', 'num_orders'
        """
        # Get limits for brand
        if brand:
            limits = self.brand_limits.get(brand, self.limits)
            brand_orders = [o for o in self.orders if o.get('Brand', '').upper() == brand.upper()]
        else:
            limits = self.limits
            brand = 'BVI'
            brand_orders = [o for o in self.orders if o.get('Brand', '').upper() == 'BVI']
        
        if not brand_orders:
            return []
        
        hours_limit = limits['Hours']
        picks_limit = limits['Picks']
        qty_limit = limits['Qty']
        offline_limit = limits.get('Offline Jobs', float('inf'))
        
        # Prepare orders with metrics including difficulty scores
        orders_with_metrics = []
        for order in brand_orders:
            qty = order.get('Lot Size', 0) or 0
            picks = order.get('Picks', 0) or 0
            hours = order.get('Hours', 0) or 0
            
            if qty > 0 and hours > 0:
                difficulty_score = self._calculate_difficulty_score(order)
                orders_with_metrics.append({
                    'order': order,
                    'qty': qty,
                    'picks': picks,
                    'hours': hours,
                    'start_date': order.get('Start Date') or datetime.max,
                    'line': self._get_line_category(order.get('Suggested Line', '')),
                    'difficulty_score': difficulty_score,
                })
        
        # Calculate difficulty thresholds and classify orders
        difficulty_thresholds = self._calculate_difficulty_thresholds(orders_with_metrics)
        for item in orders_with_metrics:
            item['difficulty'] = self._classify_difficulty(item['difficulty_score'], difficulty_thresholds)
        
        # Calculate totals
        total_hours = sum(item['hours'] for item in orders_with_metrics)
        total_picks = sum(item['picks'] for item in orders_with_metrics)
        total_qty = sum(item['qty'] for item in orders_with_metrics)
        total_orders = len(orders_with_metrics)
        
        # Calculate line distribution in source data
        line_counts_source = {'C1': 0, 'C2': 0, 'C3/4': 0, 'Other': 0}
        for item in orders_with_metrics:
            line_counts_source[item['line']] += 1
        
        total_c1_c2_c34 = line_counts_source['C1'] + line_counts_source['C2'] + line_counts_source['C3/4']
        if total_c1_c2_c34 > 0:
            line_ratios = {
                'C1': line_counts_source['C1'] / total_c1_c2_c34,
                'C2': line_counts_source['C2'] / total_c1_c2_c34,
                'C3/4': line_counts_source['C3/4'] / total_c1_c2_c34,
            }
        else:
            line_ratios = {'C1': 0.33, 'C2': 0.33, 'C3/4': 0.34}
        
        # Calculate actual number of full days possible
        full_days_possible = int(total_hours / hours_limit)
        remainder_hours = total_hours - (full_days_possible * hours_limit)
        actual_days = full_days_possible + (1 if remainder_hours > 0 else 0)
        num_days = min(num_days, actual_days)
        
        # Calculate target orders per day
        target_orders_per_day = total_orders / num_days if num_days > 0 else total_orders
        
        print(f"\n{'='*60}")
        print(f"MULTI-ROUND LEVELING PLANNING: {brand}")
        print(f"{'='*60}")
        print(f"Total orders: {total_orders}")
        print(f"Total hours: {total_hours:.1f}")
        print(f"Hours limit per day: {hours_limit}")
        print(f"Full days possible: {full_days_possible} at 100%")
        print(f"Remainder: {remainder_hours:.1f} hours ({remainder_hours/hours_limit*100:.1f}%)")
        print(f"Target orders per day: {target_orders_per_day:.1f}")
        print(f"\nSource line distribution:")
        print(f"  C1: {line_counts_source['C1']} ({line_ratios['C1']*100:.1f}%)")
        print(f"  C2: {line_counts_source['C2']} ({line_ratios['C2']*100:.1f}%)")
        print(f"  C3/4: {line_counts_source['C3/4']} ({line_ratios['C3/4']*100:.1f}%)")
        print(f"  Other: {line_counts_source['Other']}")
        
        # Print difficulty distribution
        easy_count = sum(1 for item in orders_with_metrics if item['difficulty'] == 'Easy')
        medium_count = sum(1 for item in orders_with_metrics if item['difficulty'] == 'Medium')
        hard_count = sum(1 for item in orders_with_metrics if item['difficulty'] == 'Hard')
        print(f"\nSource difficulty distribution:")
        print(f"  Easy: {easy_count} ({easy_count/total_orders*100:.1f}%)")
        print(f"  Medium: {medium_count} ({medium_count/total_orders*100:.1f}%)")
        print(f"  Hard: {hard_count} ({hard_count/total_orders*100:.1f}%)")
        print(f"  Thresholds: Easy<={difficulty_thresholds['easy']:.3f}, Hard>={difficulty_thresholds['hard']:.3f}")
        
        # Sort by start date (earlier first), then by hours (larger first for better packing)
        orders_with_metrics.sort(key=lambda x: (x['start_date'], -x['hours']))
        
        # ============================================
        # PHASE 1: Round-Robin Initial Distribution
        # ============================================
        print(f"\n--- Phase 1: Round-Robin Distribution ---")
        
        # Initialize days
        days = []
        for day_num in range(1, num_days + 1):
            days.append({
                'day': day_num,
                'orders': [],
                'items': [],  # Store the item dicts for easier manipulation
                'totals': {'Qty': 0, 'Picks': 0, 'Hours': 0},
                'num_orders': 0,
                'line_counts': {'C1': 0, 'C2': 0, 'C3/4': 0, 'Other': 0},
                'line_hours': {'C1': 0, 'C2': 0, 'C3/4': 0, 'Other': 0},
                'offline_count': 0,
                'difficulty_sum': 0.0,  # Sum of difficulty scores for averaging
                'difficulty_counts': {'Easy': 0, 'Medium': 0, 'Hard': 0},
            })
        
        # Round-robin distribution (like dealing cards)
        for idx, item in enumerate(orders_with_metrics):
            day_idx = idx % num_days
            self._add_order_to_day(days[day_idx], item)
            days[day_idx]['items'].append(item)
        
        print("After round-robin distribution:")
        for day in days:
            print(f"  Day {day['day']}: {day['num_orders']} orders, {day['totals']['Hours']:.1f} hours")
        
        # ============================================
        # PHASE 2: Hours Balancing - Maximize Days 1 to N-1
        # ============================================
        print(f"\n--- Phase 2: Hours Balancing (Maximize non-remainder days) ---")
        
        # Strategy: 
        # 1. Days 1 to N-1 should be as close to 100% as possible
        # 2. Day N (last day) is the "remainder" - can have less hours AND fewer orders
        # 3. Move orders FROM remainder day TO under-filled days
        
        remainder_day_idx = num_days - 1  # Last day is remainder
        remainder_day = days[remainder_day_idx]
        non_remainder_days = days[:remainder_day_idx]
        
        print(f"  Remainder day: Day {remainder_day['day']}")
        print(f"  Target: Maximize hours on Days 1-{remainder_day_idx}")
        
        # Phase 2a: Move orders FROM remainder TO under-filled days
        max_move_rounds = 200
        moves_made = 0
        for move_round in range(max_move_rounds):
            # Find the non-remainder day most under 100%
            under_days = [(i, d) for i, d in enumerate(non_remainder_days) 
                         if d['totals']['Hours'] < hours_limit * 0.995]
            
            if not under_days:
                print(f"  All non-remainder days at 99.5%+ after {moves_made} moves")
                break
            
            # Sort by how much under they are
            under_days.sort(key=lambda x: x[1]['totals']['Hours'])
            target_day_idx, target_day = under_days[0]
            
            hours_needed = hours_limit - target_day['totals']['Hours']
            
            if hours_needed < 1:
                break
            
            # Find best order to move from remainder to this day
            best_order = None
            best_fit_score = -float('inf')
            
            for item in remainder_day['items']:
                # How well does this order fit the gap?
                if item['hours'] <= hours_needed * 1.05:  # Allow slight overage
                    # Prefer orders that fill more of the gap
                    fit_score = item['hours'] / hours_needed if hours_needed > 0 else 0
                    # Bonus for orders that get us close to 100%
                    new_hours = target_day['totals']['Hours'] + item['hours']
                    if new_hours >= hours_limit * 0.99 and new_hours <= hours_limit * 1.02:
                        fit_score += 1.0
                    
                    if fit_score > best_fit_score:
                        # Check offline limit
                        item_line = item['order'].get('Suggested Line', '').strip().upper()
                        new_target_offline = target_day['offline_count'] + (1 if item_line == 'OFFLINE' else 0)
                        
                        if new_target_offline <= offline_limit:
                            best_fit_score = fit_score
                            best_order = item
            
            if best_order:
                # Move the order
                self._remove_order_from_day(remainder_day, best_order)
                remainder_day['items'].remove(best_order)
                
                self._add_order_to_day(target_day, best_order)
                target_day['items'].append(best_order)
                moves_made += 1
            else:
                # No single order fits - try swapping
                # Find a large order in remainder and a small order in target to swap
                best_swap = None
                best_swap_score = -float('inf')
                
                for item_remainder in remainder_day['items']:
                    for item_target in target_day['items']:
                        hours_transfer = item_remainder['hours'] - item_target['hours']
                        
                        if hours_transfer > 0:  # We want to move hours TO target
                            new_target_hours = target_day['totals']['Hours'] + hours_transfer
                            
                            # Check if this gets target closer to 100%
                            old_dev = abs(target_day['totals']['Hours'] - hours_limit)
                            new_dev = abs(new_target_hours - hours_limit)
                            
                            if new_dev < old_dev and new_target_hours <= hours_limit * 1.02:
                                improvement = old_dev - new_dev
                                
                                # Check offline limits
                                rem_line = item_remainder['order'].get('Suggested Line', '').strip().upper()
                                tgt_line = item_target['order'].get('Suggested Line', '').strip().upper()
                                
                                new_rem_offline = remainder_day['offline_count']
                                new_tgt_offline = target_day['offline_count']
                                if rem_line == 'OFFLINE':
                                    new_rem_offline -= 1
                                    new_tgt_offline += 1
                                if tgt_line == 'OFFLINE':
                                    new_tgt_offline -= 1
                                    new_rem_offline += 1
                                
                                if new_rem_offline <= offline_limit and new_tgt_offline <= offline_limit:
                                    if improvement > best_swap_score:
                                        best_swap_score = improvement
                                        best_swap = (item_remainder, item_target)
                
                if best_swap and best_swap_score > 0.1:
                    item_rem, item_tgt = best_swap
                    
                    self._remove_order_from_day(remainder_day, item_rem)
                    remainder_day['items'].remove(item_rem)
                    
                    self._remove_order_from_day(target_day, item_tgt)
                    target_day['items'].remove(item_tgt)
                    
                    self._add_order_to_day(remainder_day, item_tgt)
                    remainder_day['items'].append(item_tgt)
                    
                    self._add_order_to_day(target_day, item_rem)
                    target_day['items'].append(item_rem)
                    moves_made += 1
                else:
                    # Move to next under-filled day
                    if len(under_days) > 1:
                        continue
                    break
        
        # Phase 2b: Balance hours among non-remainder days
        print(f"  Balancing hours among Days 1-{remainder_day_idx}...")
        max_swap_rounds = 100
        for swap_round in range(max_swap_rounds):
            # Find most over and under among non-remainder days
            day_hours = [(i, d['totals']['Hours']) for i, d in enumerate(non_remainder_days)]
            day_hours.sort(key=lambda x: x[1])
            
            if len(day_hours) < 2:
                break
            
            most_under_idx = day_hours[0][0]
            most_over_idx = day_hours[-1][0]
            
            hours_spread = day_hours[-1][1] - day_hours[0][1]
            
            if hours_spread < hours_limit * 0.03:  # Within 3% spread among non-remainder
                print(f"  Non-remainder days balanced (spread: {hours_spread:.1f})")
                break
            
            day_under = non_remainder_days[most_under_idx]
            day_over = non_remainder_days[most_over_idx]
            
            # Try swapping
            best_swap = None
            best_score = -float('inf')
            
            for item_over in day_over['items']:
                for item_under in day_under['items']:
                    new_over_hours = day_over['totals']['Hours'] - item_over['hours'] + item_under['hours']
                    new_under_hours = day_under['totals']['Hours'] - item_under['hours'] + item_over['hours']
                    
                    old_over_dev = abs(day_over['totals']['Hours'] - hours_limit)
                    old_under_dev = abs(day_under['totals']['Hours'] - hours_limit)
                    new_over_dev = abs(new_over_hours - hours_limit)
                    new_under_dev = abs(new_under_hours - hours_limit)
                    
                    improvement = (old_over_dev + old_under_dev) - (new_over_dev + new_under_dev)
                    
                    if improvement > best_score:
                        over_line = item_over['order'].get('Suggested Line', '').strip().upper()
                        under_line = item_under['order'].get('Suggested Line', '').strip().upper()
                        
                        new_over_offline = day_over['offline_count']
                        new_under_offline = day_under['offline_count']
                        if over_line == 'OFFLINE':
                            new_over_offline -= 1
                            new_under_offline += 1
                        if under_line == 'OFFLINE':
                            new_under_offline -= 1
                            new_over_offline += 1
                        
                        if new_over_offline <= offline_limit and new_under_offline <= offline_limit:
                            best_score = improvement
                            best_swap = (item_over, item_under)
            
            if best_swap and best_score > 0.1:
                item_over, item_under = best_swap
                
                self._remove_order_from_day(day_over, item_over)
                day_over['items'].remove(item_over)
                
                self._remove_order_from_day(day_under, item_under)
                day_under['items'].remove(item_under)
                
                self._add_order_to_day(day_over, item_under)
                day_over['items'].append(item_under)
                
                self._add_order_to_day(day_under, item_over)
                day_under['items'].append(item_over)
            else:
                break
        
        print("\nAfter hours balancing:")
        for day in days:
            pct = day['totals']['Hours'] / hours_limit * 100
            remainder_tag = " (REMAINDER)" if day['day'] == num_days else ""
            print(f"  Day {day['day']}: {day['num_orders']} orders, {day['totals']['Hours']:.1f} hours ({pct:.1f}%){remainder_tag}")
        
        # ============================================
        # PHASE 3: Order Count Leveling
        # ============================================
        print(f"\n--- Phase 3: Order Count Leveling ---")
        
        # If order counts are very uneven, try to level them while maintaining hours
        max_order_rounds = 30
        for order_round in range(max_order_rounds):
            order_counts = [d['num_orders'] for d in days]
            order_spread = max(order_counts) - min(order_counts)
            
            if order_spread <= 5:  # Acceptable spread
                print(f"  Order count spread {order_spread} is acceptable, stopping")
                break
            
            # Find day with most and fewest orders
            max_order_day_idx = order_counts.index(max(order_counts))
            min_order_day_idx = order_counts.index(min(order_counts))
            
            day_more = days[max_order_day_idx]
            day_fewer = days[min_order_day_idx]
            
            # Try to find a swap that levels orders without hurting hours too much
            best_swap = None
            best_score = -float('inf')
            
            for item_more in day_more['items']:
                for item_fewer in day_fewer['items']:
                    # After swap, order counts stay same but check hours impact
                    new_more_hours = day_more['totals']['Hours'] - item_more['hours'] + item_fewer['hours']
                    new_fewer_hours = day_fewer['totals']['Hours'] - item_fewer['hours'] + item_more['hours']
                    
                    # Score: prefer if hours stay close to target
                    old_more_dev = abs(day_more['totals']['Hours'] - hours_limit)
                    old_fewer_dev = abs(day_fewer['totals']['Hours'] - hours_limit)
                    new_more_dev = abs(new_more_hours - hours_limit)
                    new_fewer_dev = abs(new_fewer_hours - hours_limit)
                    
                    hours_penalty = (new_more_dev + new_fewer_dev) - (old_more_dev + old_fewer_dev)
                    
                    # If this swap helps level hours even slightly, or doesn't hurt too much
                    # AND it's a swap with different hours values (to help level hours)
                    hours_diff = abs(item_more['hours'] - item_fewer['hours'])
                    
                    # We want to move hours from over to under, so:
                    # If day_more is over target and day_fewer is under, prefer item_more with high hours
                    # and item_fewer with low hours
                    if day_more['totals']['Hours'] > hours_limit and day_fewer['totals']['Hours'] < hours_limit:
                        # Want item_more.hours > item_fewer.hours to transfer hours to fewer
                        if item_more['hours'] > item_fewer['hours']:
                            score = hours_diff - hours_penalty * 2
                        else:
                            score = -hours_diff - hours_penalty * 2
                    else:
                        score = -hours_penalty
                    
                    if score > best_score:
                        # Check offline limits
                        more_line = item_more['order'].get('Suggested Line', '').strip().upper()
                        fewer_line = item_fewer['order'].get('Suggested Line', '').strip().upper()
                        
                        new_more_offline = day_more['offline_count']
                        new_fewer_offline = day_fewer['offline_count']
                        if more_line == 'OFFLINE':
                            new_more_offline -= 1
                            new_fewer_offline += 1
                        if fewer_line == 'OFFLINE':
                            new_fewer_offline -= 1
                            new_more_offline += 1
                        
                        if new_more_offline <= offline_limit and new_fewer_offline <= offline_limit:
                            best_score = score
                            best_swap = (item_more, item_fewer)
            
            if best_swap:
                item_more, item_fewer = best_swap
                
                # Perform the swap
                self._remove_order_from_day(day_more, item_more)
                day_more['items'].remove(item_more)
                
                self._remove_order_from_day(day_fewer, item_fewer)
                day_fewer['items'].remove(item_fewer)
                
                self._add_order_to_day(day_more, item_fewer)
                day_more['items'].append(item_fewer)
                
                self._add_order_to_day(day_fewer, item_more)
                day_fewer['items'].append(item_more)
            else:
                # Try moving an order from the day with more to the day with fewer
                # Find a small-hours order to move (minimal hours impact)
                day_more_items_sorted = sorted(day_more['items'], key=lambda x: x['hours'])
                
                moved = False
                for item in day_more_items_sorted[:5]:  # Try smallest 5
                    new_more_hours = day_more['totals']['Hours'] - item['hours']
                    new_fewer_hours = day_fewer['totals']['Hours'] + item['hours']
                    
                    # Only move if it doesn't hurt hours balance too much
                    old_more_dev = abs(day_more['totals']['Hours'] - hours_limit)
                    old_fewer_dev = abs(day_fewer['totals']['Hours'] - hours_limit)
                    new_more_dev = abs(new_more_hours - hours_limit)
                    new_fewer_dev = abs(new_fewer_hours - hours_limit)
                    
                    hours_penalty = (new_more_dev + new_fewer_dev) - (old_more_dev + old_fewer_dev)
                    
                    # Allow move if hours penalty is small or if it helps hours
                    if hours_penalty < item['hours'] * 0.5:
                        item_line = item['order'].get('Suggested Line', '').strip().upper()
                        new_fewer_offline = day_fewer['offline_count'] + (1 if item_line == 'OFFLINE' else 0)
                        
                        if new_fewer_offline <= offline_limit:
                            self._remove_order_from_day(day_more, item)
                            day_more['items'].remove(item)
                            
                            self._add_order_to_day(day_fewer, item)
                            day_fewer['items'].append(item)
                            moved = True
                            break
                
                if not moved:
                    print(f"  Round {order_round + 1}: No viable move/swap found, stopping")
                    break
        
        print("\nAfter order count leveling:")
        for day in days:
            print(f"  Day {day['day']}: {day['num_orders']} orders, {day['totals']['Hours']:.1f} hours")
        
        # ============================================
        # PHASE 4: Difficulty Balancing
        # ============================================
        # Goal: Each day should have similar AVERAGE difficulty
        # Swap Hard orders for Easy orders between days to balance
        print(f"\n--- Phase 4: Difficulty Balancing ---")
        
        # Calculate average difficulty across all orders
        total_difficulty = sum(d['difficulty_sum'] for d in days)
        target_avg_difficulty = total_difficulty / total_orders if total_orders > 0 else 1.0
        
        print(f"  Target avg difficulty per order: {target_avg_difficulty:.3f}")
        
        max_difficulty_rounds = 30
        for diff_round in range(max_difficulty_rounds):
            # Calculate average difficulty for each day
            day_avg_difficulties = []
            for d in days:
                if d['num_orders'] > 0:
                    avg = d['difficulty_sum'] / d['num_orders']
                else:
                    avg = target_avg_difficulty
                day_avg_difficulties.append(avg)
            
            # Find most and least difficult days
            min_diff_idx = day_avg_difficulties.index(min(day_avg_difficulties))
            max_diff_idx = day_avg_difficulties.index(max(day_avg_difficulties))
            
            difficulty_spread = max(day_avg_difficulties) - min(day_avg_difficulties)
            
            if difficulty_spread < target_avg_difficulty * 0.15:  # Within 15% spread is acceptable
                print(f"  Round {diff_round + 1}: Difficulty spread {difficulty_spread:.3f} is acceptable, stopping")
                break
            
            # The day with highest avg difficulty needs easier orders
            # The day with lowest avg difficulty can take harder orders
            hard_day = days[max_diff_idx]
            easy_day = days[min_diff_idx]
            
            # Find best swap: move a hard order from hard_day to easy_day
            # and move an easy order from easy_day to hard_day
            best_swap = None
            best_improvement = 0
            
            for item_hard in hard_day['items']:
                if item_hard.get('difficulty') != 'Hard':
                    continue  # Only consider moving hard orders out
                    
                for item_easy in easy_day['items']:
                    if item_easy.get('difficulty') != 'Easy':
                        continue  # Only swap with easy orders
                    
                    # Calculate what this swap would do to hours
                    new_hard_day_hours = hard_day['totals']['Hours'] - item_hard['hours'] + item_easy['hours']
                    new_easy_day_hours = easy_day['totals']['Hours'] - item_easy['hours'] + item_hard['hours']
                    
                    # Check if hours would be too far from target
                    old_hard_dev = abs(hard_day['totals']['Hours'] - hours_limit)
                    old_easy_dev = abs(easy_day['totals']['Hours'] - hours_limit)
                    new_hard_dev = abs(new_hard_day_hours - hours_limit)
                    new_easy_dev = abs(new_easy_day_hours - hours_limit)
                    
                    hours_penalty = (new_hard_dev + new_easy_dev) - (old_hard_dev + old_easy_dev)
                    
                    # Don't accept swaps that hurt hours too much
                    if hours_penalty > hours_limit * 0.05:
                        continue
                    
                    # Calculate difficulty improvement
                    diff_change = item_hard['difficulty_score'] - item_easy['difficulty_score']
                    
                    # After swap:
                    # hard_day loses item_hard, gains item_easy -> lower avg (good)
                    # easy_day loses item_easy, gains item_hard -> higher avg (acceptable)
                    
                    # Score the improvement
                    improvement = diff_change - hours_penalty * 0.1
                    
                    if improvement > best_improvement:
                        # Check offline limits
                        hard_line = item_hard['order'].get('Suggested Line', '').strip().upper()
                        easy_line = item_easy['order'].get('Suggested Line', '').strip().upper()
                        
                        new_hard_offline = hard_day['offline_count']
                        new_easy_offline = easy_day['offline_count']
                        if hard_line == 'OFFLINE':
                            new_hard_offline -= 1
                            new_easy_offline += 1
                        if easy_line == 'OFFLINE':
                            new_easy_offline -= 1
                            new_hard_offline += 1
                        
                        if new_hard_offline <= offline_limit and new_easy_offline <= offline_limit:
                            best_improvement = improvement
                            best_swap = (item_hard, item_easy)
            
            if best_swap and best_improvement > 0.05:
                item_hard, item_easy = best_swap
                
                # Perform the swap
                self._remove_order_from_day(hard_day, item_hard)
                hard_day['items'].remove(item_hard)
                
                self._remove_order_from_day(easy_day, item_easy)
                easy_day['items'].remove(item_easy)
                
                self._add_order_to_day(hard_day, item_easy)
                hard_day['items'].append(item_easy)
                
                self._add_order_to_day(easy_day, item_hard)
                easy_day['items'].append(item_hard)
                
                if (diff_round + 1) % 10 == 0:
                    print(f"  Round {diff_round + 1}: Swapped Hard<->Easy, improvement={best_improvement:.3f}")
            else:
                # No good swap found
                if diff_round == 0:
                    print(f"  No beneficial difficulty swaps found")
                break
        
        # Print difficulty status
        print("\nAfter difficulty balancing:")
        for day in days:
            if day['num_orders'] > 0:
                avg_diff = day['difficulty_sum'] / day['num_orders']
                counts = day['difficulty_counts']
                print(f"  Day {day['day']}: avg_diff={avg_diff:.3f}, "
                      f"Easy={counts['Easy']}, Med={counts['Medium']}, Hard={counts['Hard']}")
        
        # ============================================
        # PHASE 5: Final Cleanup
        # ============================================
        
        # Remove items list (no longer needed) and finalize
        for day in days:
            if 'items' in day:
                del day['items']
            
            day['utilization'] = {
                'Qty': day['totals']['Qty'] / qty_limit * 100 if qty_limit > 0 else 0,
                'Picks': day['totals']['Picks'] / picks_limit * 100 if picks_limit > 0 else 0,
                'Hours': day['totals']['Hours'] / hours_limit * 100 if hours_limit > 0 else 0
            }
            day['brand'] = brand
            day['day_label'] = f"Day {day['day']}"
            day['offline_limit'] = offline_limit
            day['line_distribution'] = {
                'C1': {'count': day['line_counts']['C1'], 'hours': day['line_hours']['C1']},
                'C2': {'count': day['line_counts']['C2'], 'hours': day['line_hours']['C2']},
                'C3/4': {'count': day['line_counts']['C3/4'], 'hours': day['line_hours']['C3/4']},
                'Other': {'count': day['line_counts']['Other'], 'hours': day['line_hours']['Other']}
            }
            
            # Add difficulty info
            if day['num_orders'] > 0:
                day['avg_difficulty'] = day['difficulty_sum'] / day['num_orders']
            else:
                day['avg_difficulty'] = 0
        
        # Print final summary
        print(f"\n{'='*60}")
        print("FINAL DISTRIBUTION:")
        print(f"{'='*60}")
        for day in days:
            print(f"  Day {day['day']}: {day['num_orders']} orders, "
                  f"{day['totals']['Hours']:.1f} hours ({day['utilization']['Hours']:.1f}%), "
                  f"avg_diff={day.get('avg_difficulty', 0):.3f}")
        
        return days
    
    def _remove_order_from_day(self, day: Dict, item: Dict):
        """Helper to remove an order from a day and update all tracking."""
        order = item['order']
        if order in day['orders']:
            day['orders'].remove(order)
        day['totals']['Qty'] -= item['qty']
        day['totals']['Picks'] -= item['picks']
        day['totals']['Hours'] -= item['hours']
        day['num_orders'] -= 1
        
        line = item['line']
        day['line_counts'][line] -= 1
        day['line_hours'][line] -= item['hours']
        
        order_line_raw = order.get('Suggested Line', '').strip()
        if order_line_raw.upper() == 'OFFLINE':
            day['offline_count'] -= 1
        
        # Track difficulty
        if 'difficulty_sum' in day:
            day['difficulty_sum'] -= item.get('difficulty_score', 0)
        if 'difficulty_counts' in day and 'difficulty' in item:
            day['difficulty_counts'][item['difficulty']] -= 1
    
    def _fill_day_progressive(self, day_num: int, available_orders: List[Dict], 
                              target_hours: float, hours_limit: float,
                              line_ratios: Dict, offline_limit: float, brand: str,
                              is_last_day: bool = False) -> Dict:
        """
        Fill a single day to target hours with proportional line balance.
        
        Strategy:
        1. Calculate how many orders from each line we want (proportional)
        2. Greedily select orders to hit hours target while respecting proportions
        3. Prioritize earlier start dates
        """
        day = {
            'day': day_num,
            'orders': [],
            'totals': {'Qty': 0, 'Picks': 0, 'Hours': 0},
            'num_orders': 0,
            'line_counts': {'C1': 0, 'C2': 0, 'C3/4': 0, 'Other': 0},
            'line_hours': {'C1': 0, 'C2': 0, 'C3/4': 0, 'Other': 0},
            'offline_count': 0,
        }
        
        # Group available orders by line
        orders_by_line = {'C1': [], 'C2': [], 'C3/4': [], 'Other': []}
        for item in available_orders:
            orders_by_line[item['line']].append(item)
        
        # Calculate target counts per line based on proportions and expected order count
        # Estimate order count based on average hours per order
        avg_hours_per_order = sum(item['hours'] for item in available_orders) / len(available_orders) if available_orders else 10
        estimated_orders = int(target_hours / avg_hours_per_order)
        
        # Calculate target line counts (proportional)
        c1_c2_c34_target = int(estimated_orders * 0.7)  # ~70% from C1/C2/C3/4
        target_line_counts = {
            'C1': int(c1_c2_c34_target * line_ratios['C1']),
            'C2': int(c1_c2_c34_target * line_ratios['C2']),
            'C3/4': int(c1_c2_c34_target * line_ratios['C3/4']),
        }
        
        # Phase 1: Ensure at least one from each target line (for balance)
        target_lines = ['C1', 'C2', 'C3/4']
        for line in target_lines:
            if orders_by_line[line] and day['line_counts'][line] == 0:
                # Pick earliest-dated order from this line that fits
                for item in orders_by_line[line]:
                    if day['totals']['Hours'] + item['hours'] <= target_hours * 1.1:
                        self._add_order_to_day(day, item)
                        orders_by_line[line].remove(item)
                        break
        
        # Phase 2: Fill to target hours using proportional selection
        max_iterations = len(available_orders) * 2
        iteration = 0
        
        while iteration < max_iterations:
            iteration += 1
            
            current_hours = day['totals']['Hours']
            hours_remaining = target_hours - current_hours
            
            # Stop if we've hit target (within 1%)
            if current_hours >= target_hours * 0.99:
                break
            
            # Stop if we're close enough and can't find good fits
            if current_hours >= target_hours * 0.95 and hours_remaining < 5:
                break
            
            # Find best order to add
            best_order = None
            best_score = -float('inf')
            best_line = None
            
            for line in ['C1', 'C2', 'C3/4', 'Other']:
                for item in orders_by_line[line]:
                    hours = item['hours']
                    new_hours = current_hours + hours
                    
                    # Don't exceed target by more than 2%
                    if new_hours > target_hours * 1.02:
                        continue
                    
                    # Don't exceed hard limit by more than 5%
                    if new_hours > hours_limit * 1.05:
                        continue
                    
                    # Check offline limit
                    order_line_raw = item['order'].get('Suggested Line', '').strip()
                    is_offline = order_line_raw.upper() == 'OFFLINE'
                    if is_offline and day['offline_count'] >= offline_limit:
                        continue
                    
                    # SCORING
                    score = 0
                    
                    # 1. Hours fit score (how well does this fill remaining hours?)
                    if hours <= hours_remaining:
                        # Fits within remaining - prefer orders that fill more
                        fill_ratio = hours / hours_remaining if hours_remaining > 0 else 0
                        score += fill_ratio * 50
                        # Bonus for good fits (50-100% of remaining)
                        if 0.5 <= fill_ratio <= 1.0:
                            score += 30
                    else:
                        # Would overshoot - penalty based on overage
                        overage = hours - hours_remaining
                        score -= overage * 5
                    
                    # 2. Date priority (earlier = better)
                    # Calculate days from earliest available
                    earliest = min(i['start_date'] for i in available_orders)
                    days_from_earliest = (item['start_date'] - earliest).days if earliest != datetime.max else 0
                    date_score = max(0, 30 - days_from_earliest)  # Up to 30 points for early dates
                    score += date_score
                    
                    # 3. Line balance score
                    if line in ['C1', 'C2', 'C3/4']:
                        current_line_count = day['line_counts'][line]
                        target_count = target_line_counts.get(line, 0)
                        
                        if current_line_count < target_count:
                            # Under target for this line - bonus
                            score += 20
                        elif current_line_count >= target_count * 1.5:
                            # Over target - penalty
                            score -= 15
                    
                    if score > best_score:
                        best_score = score
                        best_order = item
                        best_line = line
            
            if best_order:
                self._add_order_to_day(day, best_order)
                orders_by_line[best_line].remove(best_order)
            else:
                # No suitable order found - try to find ANY order that fits
                found = False
                for line in ['C1', 'C2', 'C3/4', 'Other']:
                    for item in orders_by_line[line]:
                        if day['totals']['Hours'] + item['hours'] <= target_hours * 1.05:
                            order_line_raw = item['order'].get('Suggested Line', '').strip()
                            is_offline = order_line_raw.upper() == 'OFFLINE'
                            if not (is_offline and day['offline_count'] >= offline_limit):
                                self._add_order_to_day(day, item)
                                orders_by_line[line].remove(item)
                                found = True
                                break
                    if found:
                        break
                
                if not found:
                    break  # Can't add any more orders
        
        # Phase 3: Fine-tune to get closer to exactly 100%
        if day['totals']['Hours'] < target_hours * 0.98:
            # Try adding small orders to fill the gap
            all_remaining = []
            for line in orders_by_line:
                all_remaining.extend(orders_by_line[line])
            
            for item in sorted(all_remaining, key=lambda x: x['hours']):
                if day['totals']['Hours'] + item['hours'] <= target_hours * 1.02:
                    order_line_raw = item['order'].get('Suggested Line', '').strip()
                    is_offline = order_line_raw.upper() == 'OFFLINE'
                    if not (is_offline and day['offline_count'] >= offline_limit):
                        self._add_order_to_day(day, item)
                        orders_by_line[item['line']].remove(item)
                        if day['totals']['Hours'] >= target_hours * 0.99:
                            break
        
        # Set line distribution
        day['line_distribution'] = {
            'C1': {'count': day['line_counts']['C1'], 'hours': day['line_hours']['C1']},
            'C2': {'count': day['line_counts']['C2'], 'hours': day['line_hours']['C2']},
            'C3/4': {'count': day['line_counts']['C3/4'], 'hours': day['line_hours']['C3/4']},
            'Other': {'count': day['line_counts']['Other'], 'hours': day['line_hours']['Other']}
        }
        
        return day
    
    def _add_order_to_day(self, day: Dict, item: Dict):
        """Helper to add an order to a day and update all tracking."""
        order = item['order']
        day['orders'].append(order)
        day['totals']['Qty'] += item['qty']
        day['totals']['Picks'] += item['picks']
        day['totals']['Hours'] += item['hours']
        day['num_orders'] += 1
        
        line = item['line']
        day['line_counts'][line] += 1
        day['line_hours'][line] += item['hours']
        
        order_line_raw = order.get('Suggested Line', '').strip()
        if order_line_raw.upper() == 'OFFLINE':
            day['offline_count'] += 1
        
        # Track difficulty
        if 'difficulty_sum' in day:
            day['difficulty_sum'] += item.get('difficulty_score', 0)
        if 'difficulty_counts' in day and 'difficulty' in item:
            day['difficulty_counts'][item['difficulty']] += 1
    
    def _create_remainder(self, remaining_orders: List[Dict], hours_limit: float,
                         qty_limit: float, picks_limit: float, 
                         offline_limit: float, brand: str) -> Dict:
        """Create a remainder day with leftover orders."""
        remainder = {
            'day': 'Remainder',
            'orders': [item['order'] for item in remaining_orders],
            'totals': {
                'Qty': sum(item['qty'] for item in remaining_orders),
                'Picks': sum(item['picks'] for item in remaining_orders),
                'Hours': sum(item['hours'] for item in remaining_orders)
            },
            'utilization': {
                'Qty': sum(item['qty'] for item in remaining_orders) / qty_limit * 100 if qty_limit > 0 else 0,
                'Picks': sum(item['picks'] for item in remaining_orders) / picks_limit * 100 if picks_limit > 0 else 0,
                'Hours': sum(item['hours'] for item in remaining_orders) / hours_limit * 100 if hours_limit > 0 else 0
            },
            'num_orders': len(remaining_orders),
            'line_counts': {'C1': 0, 'C2': 0, 'C3/4': 0, 'Other': 0},
            'line_hours': {'C1': 0, 'C2': 0, 'C3/4': 0, 'Other': 0},
            'offline_count': 0,
        }
        
        for item in remaining_orders:
            line = item['line']
            remainder['line_counts'][line] += 1
            remainder['line_hours'][line] += item['hours']
            order_line_raw = item['order'].get('Suggested Line', '').strip()
            if order_line_raw.upper() == 'OFFLINE':
                remainder['offline_count'] += 1
        
        remainder['line_distribution'] = {
            'C1': {'count': remainder['line_counts']['C1'], 'hours': remainder['line_hours']['C1']},
            'C2': {'count': remainder['line_counts']['C2'], 'hours': remainder['line_hours']['C2']},
            'C3/4': {'count': remainder['line_counts']['C3/4'], 'hours': remainder['line_hours']['C3/4']},
            'Other': {'count': remainder['line_counts']['Other'], 'hours': remainder['line_hours']['Other']}
        }
        remainder['offline_limit'] = offline_limit
        remainder['brand'] = brand
        remainder['day_label'] = 'Remainder'
        
        return remainder
    
    def _calculate_std(self, values: list) -> float:
        """Calculate standard deviation."""
        if len(values) < 2:
            return 0
        mean = sum(values) / len(values)
        variance = sum((x - mean) ** 2 for x in values) / len(values)
        return variance ** 0.5
    
    def export_to_excel(self, suggestions: List[Dict], output_path: str = 'daily_plan_suggestions.xlsx'):
        """Export suggestions to Excel file."""
        wb = openpyxl.Workbook()
        
        # Check if this is multi-day
        is_multi_day = len(suggestions) > 1 and 'day' in suggestions[0]
        
        if is_multi_day:
            ws = wb.active
            ws.title = "Multi-Day Plan"
            
            # Write summary for each day
            row = 1
            ws['A1'] = 'Day Summary'
            row += 1
            
            for suggestion in suggestions:
                day_label = suggestion.get('day_label', f"Day {suggestion.get('day', '?')}")
                ws[f'A{row}'] = day_label
                ws[f'B{row}'] = 'Orders'
                ws[f'C{row}'] = suggestion['num_orders']
                ws[f'D{row}'] = 'Qty'
                ws[f'E{row}'] = suggestion['totals']['Qty']
                ws[f'F{row}'] = 'Picks'
                ws[f'G{row}'] = suggestion['totals']['Picks']
                ws[f'H{row}'] = 'Hours'
                ws[f'I{row}'] = suggestion['totals']['Hours']
                ws[f'J{row}'] = 'Hours %'
                ws[f'K{row}'] = f"{suggestion['utilization']['Hours']:.1f}%"
                row += 1
            
            row += 1
            ws[f'A{row}'] = 'Orders:'
            row += 1
            
            # Collect all orders with day labels
            all_orders = []
            for suggestion in suggestions:
                day_label = suggestion.get('day_label', f"Day {suggestion.get('day', '?')}")
                for order in suggestion['orders']:
                    order_with_day = order.copy()
                    order_with_day['Day'] = day_label
                    all_orders.append(order_with_day)
            
            if all_orders:
                headers = ['Day'] + [k for k in all_orders[0].keys() if k != 'Day']
                for col_idx, header in enumerate(headers, 1):
                    ws.cell(row=row, column=col_idx, value=header)
                row += 1
                
                for order in all_orders:
                    col_idx = 1
                    ws.cell(row=row, column=col_idx, value=order.get('Day', ''))
                    col_idx += 1
                    for header in headers[1:]:
                        value = order.get(header, '')
                        if isinstance(value, datetime):
                            value = value.strftime('%Y-%m-%d')
                        ws.cell(row=row, column=col_idx, value=value)
                        col_idx += 1
                    row += 1
        
        wb.save(output_path)
        print(f"Exported to {output_path}")


def main():
    """Main function to run the progressive optimizer."""
    template_path = "Daily Planning Template.xlsm"
    
    optimizer = DailyPlanOptimizerProgressive(template_path=template_path)
    optimizer.load_data()
    
    print("\n" + "="*60)
    print("PROGRESSIVE FILL OPTIMIZER")
    print("="*60)
    print("\nThis optimizer will:")
    print("  1. Fill each day to 100% hours before moving to next")
    print("  2. Balance lines proportionally within each day")
    print("  3. Prioritize earlier start dates")
    print("  4. Last day gets remainder (may be < 100%)")
    
    for brand in ['BVI', 'Malosa']:
        if brand in optimizer.brand_limits:
            limits = optimizer.brand_limits[brand]
            brand_orders = [o for o in optimizer.orders if o.get('Brand', '').upper() == brand.upper()]
            
            if not brand_orders:
                print(f"\nNo orders found for {brand}")
                continue
            
            total_hours = sum(o.get('Hours', 0) or 0 for o in brand_orders)
            hours_per_day = limits['Hours']
            estimated_max_days = max(1, int(total_hours / hours_per_day) + 2)
            
            # Run the progressive multi-day planning
            day_plans = optimizer.generate_multi_day_plans(estimated_max_days, brand=brand)
            
            if day_plans:
                complete_days = [d for d in day_plans if d.get('day') != 'Remainder']
                remainder_days = [d for d in day_plans if d.get('day') == 'Remainder']
                
                print(f"\n{'='*60}")
                print(f"RESULTS: {brand}")
                print(f"{'='*60}")
                
                # Summary table
                print(f"\n{'Day':<12} {'Orders':<8} {'Qty':<10} {'Picks':<10} {'Hours':<10} {'Hours %':<10}")
                print("-" * 70)
                
                for day_plan in day_plans:
                    day_label = day_plan.get('day_label', f"Day {day_plan.get('day', '?')}")
                    print(f"{day_label:<12} {day_plan['num_orders']:<8} "
                          f"{day_plan['totals']['Qty']:<10.0f} "
                          f"{day_plan['totals']['Picks']:<10.0f} "
                          f"{day_plan['totals']['Hours']:<10.1f} "
                          f"{day_plan['utilization']['Hours']:<10.1f}%")
                
                # Calculate totals
                total_orders = sum(d['num_orders'] for d in day_plans)
                total_qty = sum(d['totals']['Qty'] for d in day_plans)
                total_picks = sum(d['totals']['Picks'] for d in day_plans)
                total_hours_planned = sum(d['totals']['Hours'] for d in day_plans)
                
                print("-" * 70)
                print(f"{'TOTAL':<12} {total_orders:<8} "
                      f"{total_qty:<10.0f} "
                      f"{total_picks:<10.0f} "
                      f"{total_hours_planned:<10.1f}")
                
                # Balance metrics for complete days only
                if len(complete_days) > 1:
                    hours_utils = [d['utilization']['Hours'] for d in complete_days]
                    orders_counts = [d['num_orders'] for d in complete_days]
                    
                    print(f"\nBalance Metrics (Complete Days, excluding Remainder):")
                    print(f"  Hours: avg={sum(hours_utils)/len(hours_utils):.1f}%, "
                          f"min={min(hours_utils):.1f}%, max={max(hours_utils):.1f}%")
                    print(f"  Orders: avg={sum(orders_counts)/len(orders_counts):.1f}, "
                          f"min={min(orders_counts)}, max={max(orders_counts)}")
                
                # Line distribution summary
                print(f"\nLine Distribution per Day:")
                for day_plan in day_plans:
                    day_label = day_plan.get('day_label', f"Day {day_plan.get('day', '?')}")
                    line_dist = day_plan.get('line_distribution', {})
                    c1 = line_dist.get('C1', {}).get('count', 0)
                    c2 = line_dist.get('C2', {}).get('count', 0)
                    c34 = line_dist.get('C3/4', {}).get('count', 0)
                    other = line_dist.get('Other', {}).get('count', 0)
                    print(f"  {day_label}: C1={c1}, C2={c2}, C3/4={c34}, Other={other}")
                
                # Start date analysis
                print(f"\nStart Date Range per Day:")
                for day_plan in day_plans:
                    day_label = day_plan.get('day_label', f"Day {day_plan.get('day', '?')}")
                    dates = [o.get('Start Date') for o in day_plan['orders'] if o.get('Start Date')]
                    if dates:
                        min_date = min(dates)
                        max_date = max(dates)
                        print(f"  {day_label}: {min_date.strftime('%Y-%m-%d')} to {max_date.strftime('%Y-%m-%d')}")
                
                # Create output directory
                output_dir = 'output'
                if not os.path.exists(output_dir):
                    os.makedirs(output_dir)
                
                # Generate timestamp
                timestamp = datetime.now().strftime('%Y%m%d%H%M')
                
                # Export
                brand_lower = brand.lower()
                excel_filename = f'{timestamp}-{brand_lower}-progressive-plan.xlsx'
                excel_path = os.path.join(output_dir, excel_filename)
                
                optimizer.export_to_excel(day_plans, excel_path)
                print(f"\nExported to: {excel_path}")
            else:
                print(f"No plans generated for {brand}")
    
    print("\n" + "="*60)
    print("Done! Check the generated Excel files.")
    print("="*60)


if __name__ == "__main__":
    main()
