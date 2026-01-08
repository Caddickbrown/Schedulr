"""
Daily Planning Optimizer
Generates optimal daily plans balancing Qty, Picks, and Hours based on limits.
Reads directly from Excel template file.
"""
import csv
import openpyxl
import os
from datetime import datetime
from typing import List, Dict, Tuple
import json


class DailyPlanOptimizer:
    def __init__(self, template_path: str = None):
        """
        Initialize optimizer with Excel template file.
        
        Args:
            template_path: Path to Excel template file (default: 'Daily Planning Template.xlsm')
        """
        self.template_path = template_path or 'Daily Planning Template.xlsm'
        self.limits = {}
        self.orders = []
        self.brand_limits = {}
        
    def load_data(self):
        """Load orders and limits directly from Excel template."""
        self._load_limits()
        self._load_orders()
        
    def _load_limits(self):
        """Load brand-specific limits directly from Excel template."""
        if not os.path.exists(self.template_path):
            raise FileNotFoundError(f"Template file not found: {self.template_path}")
        
        wb = openpyxl.load_workbook(self.template_path, data_only=True)
        main_sheet = wb['Main']
        
        # Extract limits (row 2)
        headers = [cell.value for cell in main_sheet[1]]
        limits_row = [cell.value for cell in main_sheet[2]]
        
        limits = {}
        try:
            qty_idx = headers.index('Qty')
            picks_idx = headers.index('Picks')
            hours_idx = headers.index('Hours')
            limits = {
                'Qty': self._parse_float(limits_row[qty_idx]),
                'Picks': self._parse_float(limits_row[picks_idx]),
                'Hours': self._parse_float(limits_row[hours_idx])
            }
        except ValueError:
            print("Warning: Could not extract basic limits")
        
        # Extract brand-specific limits (BVI and Malosa)
        bvi_limits = {}
        malosa_limits = {}
        
        try:
            # BVI limits are in the main section
            bvi_limits = {
                'Qty': limits.get('Qty', 0),
                'Picks': limits.get('Picks', 0),
                'Hours': limits.get('Hours', 0),
            }
            
            # Try to get additional BVI limits
            if 'Low Picks' in headers:
                try:
                    bvi_limits['Low Picks'] = self._parse_float(limits_row[headers.index('Low Picks')])
                except (ValueError, IndexError):
                    pass
            if 'Big Picks' in headers:
                try:
                    bvi_limits['Big Picks'] = self._parse_float(limits_row[headers.index('Big Picks')])
                except (ValueError, IndexError):
                    pass
            if 'Large Orders' in headers:
                try:
                    bvi_limits['Large Orders'] = self._parse_float(limits_row[headers.index('Large Orders')])
                except (ValueError, IndexError):
                    pass
            if 'Offline Jobs' in headers:
                try:
                    bvi_limits['Offline Jobs'] = self._parse_float(limits_row[headers.index('Offline Jobs')])
                except (ValueError, IndexError):
                    pass
            
            # Find Malosa section - look for "Malosa" in headers
            if 'Malosa' in headers:
                malosa_start = headers.index('Malosa')
                # Malosa Qty, Picks, Hours should be in next few columns
                for i in range(malosa_start, len(headers)):
                    if headers[i] == 'Qty' and i + 2 < len(limits_row):
                        malosa_limits = {
                            'Qty': self._parse_float(limits_row[i]),
                            'Picks': self._parse_float(limits_row[i+1]),
                            'Hours': self._parse_float(limits_row[i+2]),
                        }
                        break
        except Exception as e:
            print(f"Warning: Could not extract brand-specific limits: {e}")
        
        # Set defaults if not found
        if not bvi_limits.get('Qty'):
            bvi_limits = {'Qty': 10544, 'Picks': 750, 'Hours': 390}
        if not malosa_limits.get('Qty'):
            malosa_limits = {'Qty': 3335, 'Picks': 130, 'Hours': 90}
        
        self.brand_limits = {
            'BVI': bvi_limits,
            'Malosa': malosa_limits
        }
        
        print(f"Loaded brand limits: {self.brand_limits}")
        # Keep legacy self.limits for backward compatibility (defaults to BVI)
        self.limits = self.brand_limits.get('BVI', {})
    
    def _load_orders(self):
        """Load orders directly from Excel template."""
        if not os.path.exists(self.template_path):
            raise FileNotFoundError(f"Template file not found: {self.template_path}")
        
        wb = openpyxl.load_workbook(self.template_path, data_only=True)
        main_sheet = wb['Main']
        
        # Extract order headers (row 11)
        order_headers = [cell.value for cell in main_sheet[11]]
        
        # Extract orders (starting row 12)
        for row_idx in range(12, main_sheet.max_row + 1):
            row = [cell.value for cell in main_sheet[row_idx]]
            if not row[0] or row[0] == 'Order No':
                continue
            
            try:
                # Build order dict from row
                row_dict = {}
                for idx, header in enumerate(order_headers):
                    if idx < len(row) and header:
                        value = row[idx]
                        # Keep datetime as datetime object (don't convert to string)
                        row_dict[header] = value
                
                # Normalize line name (C3/4, C3&4 -> C3/4)
                suggested_line = str(row_dict.get('Suggested Line', '')).strip()
                if suggested_line in ['C3/4', 'C3&4']:
                    suggested_line = 'C3/4'
                
                # Calculate efficiency metrics
                qty = self._parse_float(row_dict.get('Lot Size', 0))
                picks = self._parse_float(row_dict.get('Picks', 0))
                hours = self._parse_float(row_dict.get('Hours', 0))
                
                # Try to get from Excel, otherwise calculate
                qty_per_hr = self._parse_float(row_dict.get('Qty/Hr', 0))
                picks_per_hr = self._parse_float(row_dict.get('Picks/Hr', 0))
                picks_per_qty = self._parse_float(row_dict.get('Picks/Qty', 0))
                
                # Calculate if missing
                if qty_per_hr == 0 and hours > 0:
                    qty_per_hr = qty / hours
                if picks_per_hr == 0 and hours > 0:
                    picks_per_hr = picks / hours
                if picks_per_qty == 0 and qty > 0:
                    picks_per_qty = picks / qty
                
                # Parse date - handle both datetime objects and strings
                start_date = row_dict.get('Start Date')
                if isinstance(start_date, datetime):
                    parsed_date = start_date
                elif start_date:
                    parsed_date = self._parse_date(str(start_date))
                else:
                    parsed_date = None
                
                order = {
                    'Order No': str(row_dict.get('Order No', '')).strip(),
                    'Part No': str(row_dict.get('Part No', '')).strip(),
                    'Brand': str(row_dict.get('Brand', '')).strip(),
                    'Start Date': parsed_date,
                    'Lot Size': qty,
                    'Picks': picks,
                    'Hours': hours,
                    'Country': str(row_dict.get('Country', '')).strip(),
                    'Wrap Type': str(row_dict.get('Wrap Type', '')).strip(),
                    'CPU': self._parse_float(row_dict.get('CPU', 0)),
                    'Suggested Line': suggested_line,
                    'Qty/Hr': qty_per_hr,
                    'Picks/Hr': picks_per_hr,
                    'Picks/Qty': picks_per_qty,
                }
                
                # Only add if we have essential data
                if order['Order No'] and order['Part No'] and order['Lot Size'] > 0:
                    self.orders.append(order)
            except Exception as e:
                order_no = row[0] if row else 'unknown'
                print(f"Error processing order {order_no}: {e}")
                continue
        
        print(f"Loaded {len(self.orders)} orders")
    
    def _parse_date(self, date_str):
        """Parse date string to datetime object."""
        if not date_str or date_str.strip() == '':
            return None
        try:
            # Try different date formats
            for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%Y-%m-%d %H:%M:%S']:
                try:
                    return datetime.strptime(date_str.strip(), fmt)
                except ValueError:
                    continue
            return None
        except:
            return None
    
    def _parse_float(self, value):
        """Parse float value, handling empty strings and None."""
        if value is None or value == '' or value == '-':
            return 0.0
        try:
            return float(str(value).replace(',', ''))
        except:
            return 0.0
    
    def _get_line_category(self, line: str) -> str:
        """Normalize line name to category (C1, C2, C3/4, or Other)."""
        if not line:
            return 'Other'
        line_upper = line.upper().strip()
        if line_upper in ['C1']:
            return 'C1'
        elif line_upper in ['C2']:
            return 'C2'
        elif line_upper in ['C3/4', 'C3&4', 'C3/4']:
            return 'C3/4'
        else:
            return 'Other'
    
    def _categorize_order_difficulty(self, order: Dict, brand_orders: List[Dict] = None) -> str:
        """
        Categorize order as Easy, Medium, or Difficult based on efficiency metrics.
        
        Easy: High Qty/Hr, Low Picks/Qty (efficient, low pick density)
        Medium: Moderate efficiency metrics
        Difficult: Low Qty/Hr, High Picks/Qty (inefficient, high pick density)
        """
        qty_per_hr = order.get('Qty/Hr', 0)
        picks_per_qty = order.get('Picks/Qty', 0)
        
        if qty_per_hr == 0 or picks_per_qty == 0:
            return 'Medium'
        
        # Use provided brand_orders or all orders for this brand
        if brand_orders is None:
            brand = order.get('Brand', 'BVI')
            brand_orders = [o for o in self.orders if o.get('Brand', '').upper() == brand.upper()]
        
        if len(brand_orders) < 3:
            return 'Medium'
        
        # Get efficiency metrics for all brand orders
        qty_hr_values = [o.get('Qty/Hr', 0) for o in brand_orders if o.get('Qty/Hr', 0) > 0]
        picks_qty_values = [o.get('Picks/Qty', 0) for o in brand_orders if o.get('Picks/Qty', 0) > 0]
        
        if not qty_hr_values or not picks_qty_values:
            return 'Medium'
        
        # Calculate percentiles (33rd and 67th)
        qty_hr_sorted = sorted(qty_hr_values)
        picks_qty_sorted = sorted(picks_qty_values)
        
        qty_hr_33_idx = len(qty_hr_sorted) // 3
        qty_hr_67_idx = (len(qty_hr_sorted) * 2) // 3
        picks_qty_33_idx = len(picks_qty_sorted) // 3
        picks_qty_67_idx = (len(picks_qty_sorted) * 2) // 3
        
        qty_hr_33 = qty_hr_sorted[qty_hr_33_idx] if qty_hr_33_idx < len(qty_hr_sorted) else 0
        qty_hr_67 = qty_hr_sorted[qty_hr_67_idx] if qty_hr_67_idx < len(qty_hr_sorted) else 0
        picks_qty_33 = picks_qty_sorted[picks_qty_33_idx] if picks_qty_33_idx < len(picks_qty_sorted) else 0
        picks_qty_67 = picks_qty_sorted[picks_qty_67_idx] if picks_qty_67_idx < len(picks_qty_sorted) else 0
        
        # Categorize: Easy = high Qty/Hr AND low Picks/Qty
        # Difficult = low Qty/Hr OR high Picks/Qty
        is_easy = qty_per_hr >= qty_hr_67 and picks_per_qty <= picks_qty_67
        is_difficult = qty_per_hr <= qty_hr_33 or picks_per_qty >= picks_qty_67
        
        if is_easy:
            return 'Easy'
        elif is_difficult:
            return 'Difficult'
        else:
            return 'Medium'
    
    def _calculate_line_balance_score(self, selected_orders: List[Dict]) -> float:
        """
        Calculate how balanced the line distribution is (1:1:1 ratio).
        Returns a score where higher = more balanced.
        """
        line_counts = {'C1': 0, 'C2': 0, 'C3/4': 0, 'Other': 0}
        line_hours = {'C1': 0, 'C2': 0, 'C3/4': 0, 'Other': 0}
        
        for order in selected_orders:
            line = self._get_line_category(order.get('Suggested Line', ''))
            line_counts[line] += 1
            line_hours[line] += order.get('Hours', 0) or 0
        
        # Focus on C1, C2, C3/4 (ignore Other)
        target_lines = ['C1', 'C2', 'C3/4']
        counts = [line_counts[line] for line in target_lines]
        hours = [line_hours[line] for line in target_lines]
        
        if sum(counts) == 0:
            return 0.0
        
        # Calculate balance based on both counts and hours
        # For 1:1:1 ratio, we want equal distribution
        count_mean = sum(counts) / len(counts) if len(counts) > 0 else 0
        hours_mean = sum(hours) / len(hours) if len(hours) > 0 else 0
        
        # Calculate variance (lower = more balanced)
        count_variance = sum((c - count_mean) ** 2 for c in counts) / len(counts) if len(counts) > 0 else 0
        hours_variance = sum((h - hours_mean) ** 2 for h in hours) / len(hours) if len(hours) > 0 else 0
        
        # Score: prefer lower variance (more balanced)
        # Normalize and invert so higher score = more balanced
        count_balance = 100 - (count_variance ** 0.5) * 10
        hours_balance = 100 - (hours_variance ** 0.5) * 10
        
        return (count_balance + hours_balance) / 2
    
    def optimize_plan_balanced(self, brand: str = None, limits: Dict = None) -> Tuple[List[Dict], Dict]:
        """
        Optimize daily plan with Hours as the pivot.
        - Hours MUST be hit (target the limit)
        - Target ~40 orders
        - Balance Qty and Picks around Hours constraint
        - Prioritize by due date (earlier dates first)
        
        Args:
            brand: Brand to filter orders (BVI, Malosa, etc.)
            limits: Limits dictionary to use (defaults to self.limits)
        """
        # Filter orders by brand if specified
        orders_to_optimize = self.orders
        if brand:
            orders_to_optimize = [o for o in self.orders if o.get('Brand', '').upper() == brand.upper()]
        
        if not orders_to_optimize:
            return [], {
                'totals': {'Qty': 0, 'Picks': 0, 'Hours': 0},
                'utilization': {'Qty': 0, 'Picks': 0, 'Hours': 0},
                'num_orders': 0
            }
        
        # Use provided limits or default to self.limits
        if limits is None:
            limits = self.limits
        
        target_orders = 40
        hours_target = limits['Hours']
        
        # Filter and prepare orders with metrics
        orders_with_metrics = []
        for order in orders_to_optimize:
            qty = order.get('Lot Size', 0) or 0
            picks = order.get('Picks', 0) or 0
            hours = order.get('Hours', 0) or 0
            
            if qty > 0 and hours > 0:  # Must have quantity and hours
                # Get date priority (earlier = higher priority)
                date_priority = 0.0
                if order.get('Start Date'):
                    try:
                        earliest_date = min(o.get('Start Date') for o in orders_to_optimize if o.get('Start Date'))
                        if earliest_date:
                            days_diff = (order['Start Date'] - earliest_date).days
                            # Earlier dates get higher priority (0-1 range, normalized)
                            date_priority = 1.0 - min(days_diff / 60.0, 1.0)  # 60 day window
                    except:
                        pass
                
                # Categorize order difficulty
                difficulty = self._categorize_order_difficulty(order, orders_to_optimize)
                
                orders_with_metrics.append({
                    'order': order,
                    'qty': qty,
                    'picks': picks,
                    'hours': hours,
                    'date_priority': date_priority,
                    'start_date': order.get('Start Date') or datetime.max,
                    'difficulty': difficulty
                })
        
        # Sort primarily by due date (earlier first), then by hours contribution
        orders_with_metrics.sort(key=lambda x: (x['start_date'], -x['hours']))
        
        selected = []
        totals = {'Qty': 0, 'Picks': 0, 'Hours': 0}
        line_counts = {'C1': 0, 'C2': 0, 'C3/4': 0, 'Other': 0}
        line_hours = {'C1': 0, 'C2': 0, 'C3/4': 0, 'Other': 0}
        offline_count = 0  # Track number of offline orders
        difficulty_counts = {'Easy': 0, 'Medium': 0, 'Difficult': 0}  # Track difficulty mix
        
        # Phase 1: Fill to Hours target, prioritizing by due date and line balance
        # Hours is the PIVOT - must hit this target
        # Qty and Picks are flexible (can go over if needed to hit hours)
        remaining_orders = orders_with_metrics.copy()
        hours_tolerance = 0.02  # 2% tolerance for hours (very tight)
        qty_picks_flexibility = 1.5  # Allow Qty/Picks to go 50% over if needed (very flexible)
        
        # Phase 0: Ensure at least one order from each target line early
        target_lines = ['C1', 'C2', 'C3/4']
        for target_line in target_lines:
            if line_counts[target_line] == 0:
                # Find best order for this line
                best_for_line = None
                best_score = -float('inf')
                
                for item in remaining_orders:
                    item_line = self._get_line_category(item['order'].get('Suggested Line', ''))
                    if item_line != target_line:
                        continue
                    
                    qty = item['qty']
                    picks = item['picks']
                    hours = item['hours']
                    
                    # Check basic constraints
                    if (totals['Hours'] + hours > hours_target * 1.1 or
                        totals['Qty'] + qty > limits['Qty'] * 2.0 or
                        totals['Picks'] + picks > limits['Picks'] * 2.0):
                        continue
                    
                    # Score: prefer smaller hours to leave room for more orders
                    score = -hours + item['date_priority'] * 10
                    
                    if score > best_score:
                        best_score = score
                        best_for_line = item
                
                if best_for_line:
                    selected.append(best_for_line['order'])
                    totals['Qty'] += best_for_line['qty']
                    totals['Picks'] += best_for_line['picks']
                    totals['Hours'] += best_for_line['hours']
                    order_line = self._get_line_category(best_for_line['order'].get('Suggested Line', ''))
                    line_counts[order_line] += 1
                    line_hours[order_line] += best_for_line['hours']
                    # Update difficulty tracking
                    order_difficulty = best_for_line.get('difficulty', 'Medium')
                    difficulty_counts[order_difficulty] += 1
                    remaining_orders.remove(best_for_line)
        
        iterations = 0
        max_iterations = len(orders_with_metrics) * 3  # Safety limit
        
        while remaining_orders and iterations < max_iterations:
            iterations += 1
            
            # Check if we have at least one order from each target line
            target_lines = ['C1', 'C2', 'C3/4']
            has_all_lines = all(line_counts[line] > 0 for line in target_lines)
            
            # Stop conditions: MUST reach at least 99.5% of hours target
            # 1. We've hit hours target (at least 99.5%, max 102%) AND have enough orders AND have all lines represented
            if (totals['Hours'] >= hours_target * 0.995 and 
                totals['Hours'] <= hours_target * 1.02 and
                len(selected) >= target_orders * 0.7 and  # At least 70% of target
                has_all_lines):  # Must have at least one order from each line
                break
            
            # 2. We're at or over 100% of hours target - can stop if we have reasonable orders
            # But don't stop if we're missing a line (unless we're way over hours)
            if (totals['Hours'] >= hours_target and
                len(selected) >= target_orders * 0.6):
                if has_all_lines or totals['Hours'] > hours_target * 1.05:
                    break
            
            best_order = None
            best_score = -float('inf')
            
            for item in remaining_orders:
                qty = item['qty']
                picks = item['picks']
                hours = item['hours']
                
                # Hours is the hard constraint - must stay within tolerance
                # But allow slightly more flexibility if we need to balance lines
                hours_after = totals['Hours'] + hours
                order_line = self._get_line_category(item['order'].get('Suggested Line', ''))
                max_hours_tolerance = hours_tolerance
                
                # Allow more hours flexibility if this order helps balance lines
                if order_line in ['C1', 'C2', 'C3/4']:
                    target_lines = ['C1', 'C2', 'C3/4']
                    if line_counts[order_line] == 0:
                        # Allow up to 5% over if adding first order to a line
                        max_hours_tolerance = 0.05
                    elif sum(line_counts[line] for line in target_lines) > 0:
                        # Check if this line is underrepresented
                        line_ratios = [line_counts[line] / sum(line_counts[line] for line in target_lines) 
                                      for line in target_lines]
                        target_ratio = 1.0 / 3.0
                        if line_ratios[target_lines.index(order_line)] < target_ratio * 0.7:
                            max_hours_tolerance = 0.04  # Allow 4% over for underrepresented lines
                
                if hours_after > hours_target * (1 + max_hours_tolerance):
                    continue  # Would exceed hours limit
                
                # Qty and Picks are flexible - allow significant overage to hit hours target
                qty_would_exceed = totals['Qty'] + qty > self.limits['Qty'] * qty_picks_flexibility
                picks_would_exceed = totals['Picks'] + picks > self.limits['Picks'] * qty_picks_flexibility
                
                # Only block if BOTH would exceed AND we're not close to hours target
                if qty_would_exceed and picks_would_exceed:
                    # Both would exceed - only allow if we're close to hours target or need orders
                    if totals['Hours'] < hours_target * 0.85 and len(selected) >= target_orders * 0.9:
                        continue  # Too far from target and have enough orders
                
                # Score based on:
                # 1. How close we get to hours target (CRITICAL - highest weight)
                hours_distance_from_target = abs(hours_target - hours_after)
                current_distance = abs(hours_target - totals['Hours'])
                
                if hours_after <= hours_target:
                    # We're still under target - STRONGLY prefer getting closer
                    hours_score = (current_distance - hours_distance_from_target) * 100
                    # Extra bonus if this gets us very close
                    if hours_distance_from_target < hours_target * 0.05:
                        hours_score += 50
                else:
                    # We're over target - prefer staying as close as possible
                    hours_score = -hours_distance_from_target * 200  # Strong penalty for going over
                
                # 2. Order count bonus (STRONG preference for ~40 orders)
                order_count_bonus = 0
                if len(selected) < target_orders * 0.6:
                    order_count_bonus = 40  # Very strong bonus when well below target
                elif len(selected) < target_orders * 0.8:
                    order_count_bonus = 30  # Strong bonus when below target
                elif len(selected) < target_orders:
                    order_count_bonus = 20  # Moderate bonus when approaching target
                elif len(selected) > target_orders * 1.3:
                    order_count_bonus = -20  # Penalty for too many orders
                
                # Bonus for smaller orders when we need more order count (helps reach 40)
                if len(selected) < target_orders:
                    # Prefer orders with lower qty (helps avoid hitting qty limit too early)
                    avg_order_qty = totals['Qty'] / len(selected) if len(selected) > 0 else 0
                    if qty < avg_order_qty * 0.5:  # Order is much smaller than average
                        order_count_bonus += 15  # Extra bonus for small orders
            
                # 3. Qty/Picks constraint penalty (prefer staying within limits)
                constraint_penalty = 0
                if qty_would_exceed:
                    constraint_penalty -= 5  # Small penalty for exceeding Qty
                if picks_would_exceed:
                    constraint_penalty -= 5  # Small penalty for exceeding Picks
                
                # 4. Balance score: prefer orders that help balance qty and picks
                current_qty_util = totals['Qty'] / limits['Qty'] if limits['Qty'] > 0 else 0
                current_picks_util = totals['Picks'] / limits['Picks'] if limits['Picks'] > 0 else 0
                new_qty_util = (totals['Qty'] + qty) / limits['Qty'] if limits['Qty'] > 0 else 0
                new_picks_util = (totals['Picks'] + picks) / limits['Picks'] if limits['Picks'] > 0 else 0
                
                # Calculate how balanced qty and picks are
                current_balance = abs(current_qty_util - current_picks_util)
                new_balance = abs(new_qty_util - new_picks_util)
                balance_improvement = (current_balance - new_balance) * 2
                
                # 5. Date priority bonus (earlier dates preferred)
                date_bonus = item['date_priority'] * 5
                
                # 6. Difficulty blending bonus (prefer mixing Easy/Medium/Difficult orders)
                difficulty_bonus = 0
                order_difficulty = item.get('difficulty', 'Medium')
                total_selected = len(selected)
                
                if total_selected > 0:
                    # Calculate current difficulty distribution
                    easy_ratio = difficulty_counts['Easy'] / total_selected
                    medium_ratio = difficulty_counts['Medium'] / total_selected
                    difficult_ratio = difficulty_counts['Difficult'] / total_selected
                    
                    # Target: ~30% Easy, ~40% Medium, ~30% Difficult (balanced mix)
                    target_easy = 0.30
                    target_medium = 0.40
                    target_difficult = 0.30
                    
                    # Calculate how much this order would improve the mix
                    if order_difficulty == 'Easy':
                        # Bonus if we have too few easy orders
                        if easy_ratio < target_easy * 0.8:
                            difficulty_bonus = 15 * (target_easy - easy_ratio) / target_easy
                        # Small penalty if we have too many easy orders
                        elif easy_ratio > target_easy * 1.2:
                            difficulty_bonus = -5
                    elif order_difficulty == 'Medium':
                        # Always prefer medium orders (they're the "sweet spot")
                        if medium_ratio < target_medium:
                            difficulty_bonus = 20 * (target_medium - medium_ratio) / target_medium
                    elif order_difficulty == 'Difficult':
                        # Strong bonus if we have too few difficult orders (need to blend them in)
                        if difficult_ratio < target_difficult * 0.7:
                            difficulty_bonus = 25 * (target_difficult - difficult_ratio) / target_difficult
                        # Small penalty if we have too many difficult orders
                        elif difficult_ratio > target_difficult * 1.3:
                            difficulty_bonus = -3
                else:
                    # First order - prefer medium difficulty to start balanced
                    if order_difficulty == 'Medium':
                        difficulty_bonus = 10
                
                # 7. Line balance bonus (prefer orders that help achieve 1:1:1 ratio for C1:C2:C3/4)
                # Note: order_line already calculated above
                line_balance_bonus = 0
                
                if order_line in ['C1', 'C2', 'C3/4']:
                    target_lines = ['C1', 'C2', 'C3/4']
                    current_counts = [line_counts[line] for line in target_lines]
                    
                    # CRITICAL: If a line has zero orders, strongly prefer adding to it
                    if line_counts[order_line] == 0:
                        # Very strong bonus for adding first order to a line
                        line_balance_bonus += 50
                    
                    # Calculate what the balance would be if we add this order
                    temp_line_counts = line_counts.copy()
                    temp_line_hours = line_hours.copy()
                    temp_line_counts[order_line] += 1
                    temp_line_hours[order_line] += hours
                    
                    current_hours = [line_hours[line] for line in target_lines]
                    new_counts = [temp_line_counts[line] for line in target_lines]
                    new_hours = [temp_line_hours[line] for line in target_lines]
                    
                    # Calculate how balanced we are (1:1:1 ratio)
                    # Lower variance = more balanced
                    if sum(current_counts) > 0:
                        current_mean = sum(current_counts) / len(current_counts)
                        current_variance = sum((c - current_mean) ** 2 for c in current_counts) / len(current_counts)
                    else:
                        current_variance = float('inf')
                    
                    if sum(new_counts) > 0:
                        new_mean = sum(new_counts) / len(new_counts)
                        new_variance = sum((c - new_mean) ** 2 for c in new_counts) / len(new_counts)
                    else:
                        new_variance = 0
                    
                    # Prefer orders that reduce variance (improve balance)
                    variance_improvement = current_variance - new_variance
                    
                    # Also consider hours balance
                    if sum(current_hours) > 0:
                        current_hours_mean = sum(current_hours) / len(current_hours)
                        current_hours_variance = sum((h - current_hours_mean) ** 2 for h in current_hours) / len(current_hours)
                    else:
                        current_hours_variance = float('inf')
                    
                    if sum(new_hours) > 0:
                        new_hours_mean = sum(new_hours) / len(new_hours)
                        new_hours_variance = sum((h - new_hours_mean) ** 2 for h in new_hours) / len(new_hours)
                    else:
                        new_hours_variance = 0
                    
                    hours_variance_improvement = current_hours_variance - new_hours_variance
                    
                    # Strong bonus for improving balance (weighted by how far we are from target)
                    line_balance_bonus += (variance_improvement * 40 + hours_variance_improvement * 20)
                    
                    # Extra bonus if this line is underrepresented (stronger preference)
                    if sum(new_counts) > 0:
                        line_ratio = new_counts[target_lines.index(order_line)] / sum(new_counts)
                        target_ratio = 1.0 / 3.0  # 1:1:1 ratio
                        if line_ratio < target_ratio * 0.8:  # Line is underrepresented
                            # Stronger bonus the more underrepresented
                            underrepresentation = (target_ratio - line_ratio) / target_ratio
                            line_balance_bonus += 30 * underrepresentation
                
                score = hours_score + balance_improvement + date_bonus + order_count_bonus + constraint_penalty + difficulty_bonus + line_balance_bonus
                
                if score > best_score:
                    best_score = score
                    best_order = item
            
            if best_order:
                selected.append(best_order['order'])
                totals['Qty'] += best_order['qty']
                totals['Picks'] += best_order['picks']
                totals['Hours'] += best_order['hours']
                
                # Update line tracking
                order_line = self._get_line_category(best_order['order'].get('Suggested Line', ''))
                line_counts[order_line] += 1
                line_hours[order_line] += best_order['hours']
                
                # Update difficulty tracking
                order_difficulty = best_order.get('difficulty', 'Medium')
                difficulty_counts[order_difficulty] += 1
                
                # Update offline count
                order_line_raw = best_order['order'].get('Suggested Line', '').strip()
                if order_line_raw.upper() == 'OFFLINE':
                    offline_count += 1
                
                remaining_orders.remove(best_order)
            else:
                # No more orders can fit - try relaxing constraints slightly
                if totals['Hours'] < hours_target * 0.995:  # Must reach at least 99.5%
                    # Still far from target - allow more flexibility
                    qty_picks_flexibility = min(qty_picks_flexibility * 1.1, 1.5)  # Allow up to 50% over
                    hours_tolerance = min(hours_tolerance * 1.2, 0.05)  # Allow up to 5% over
                elif totals['Hours'] >= hours_target * 0.995:
                    # We've reached at least 99.5% - this is acceptable
                    break
                else:
                    # Can't add more and we're below 99.5% - keep trying with relaxed constraints
                    qty_picks_flexibility = min(qty_picks_flexibility * 1.1, 1.5)
                    hours_tolerance = min(hours_tolerance * 1.2, 0.05)
        
        # Phase 2: Fine-tune to get closer to hours target AND order count target
        # ALWAYS run Phase 2 to try to reach 100% hours target
        # Continue adding orders if we need more orders OR need to reach 100% hours
        if len(selected) < target_orders * 1.1 or totals['Hours'] < hours_target * 0.995:
            # Sort remaining by: 1) date priority, 2) how close they get us to hours target
            remaining_orders.sort(key=lambda x: (
                x['start_date'],
                abs(hours_target - totals['Hours'] - x['hours'])
            ))
            
            for item in remaining_orders[:300]:  # Check top 300 candidates
                # Stop if we've hit hours target (at least 99.5%) and have enough orders
                if (totals['Hours'] >= hours_target * 0.995 and
                    len(selected) >= target_orders * 0.8):
                    break
                
                # Don't go too far over on orders
                if len(selected) >= target_orders * 1.5:
                    break
                    
                qty = item['qty']
                picks = item['picks']
                hours = item['hours']
                
                # Hours must stay within tolerance (max 2% over, but prefer reaching 100%)
                new_hours = totals['Hours'] + hours
                if new_hours > hours_target * 1.02:
                    continue
                
                # If we're still under target, strongly prefer orders that get us closer
                if totals['Hours'] < hours_target * 0.995:
                    # Must get closer to target
                    current_hours_distance = abs(hours_target - totals['Hours'])
                    new_hours_distance = abs(hours_target - new_hours)
                    if new_hours_distance >= current_hours_distance:
                        continue  # Skip if it doesn't improve
                
                # Qty/Picks can be flexible (up to 50% over)
                if (totals['Qty'] + qty > limits['Qty'] * 1.5 or
                    totals['Picks'] + picks > limits['Picks'] * 1.5):
                    continue
                
                # Decision logic:
                # Priority 1: Get hours as close to target as possible
                # Priority 2: Get order count to ~40
                
                current_hours_distance = abs(hours_target - totals['Hours'])
                new_hours_distance = abs(hours_target - new_hours)
                hours_improves = new_hours_distance < current_hours_distance
                
                if totals['Hours'] < hours_target * 0.995:
                    # Under 99.5% target - MUST add orders that get us closer
                    if hours_improves:
                        selected.append(item['order'])
                        totals['Qty'] += qty
                        totals['Picks'] += picks
                        totals['Hours'] += hours
                        order_difficulty = item.get('difficulty', 'Medium')
                        difficulty_counts[order_difficulty] += 1
                        remaining_orders.remove(item)
                elif totals['Hours'] >= hours_target * 0.995 and totals['Hours'] <= hours_target * 1.01:
                    # Very close to target (within 1%) - can add small orders for order count
                    if len(selected) < target_orders:
                        # Prefer small orders that don't push us too far over
                        if hours < hours_target * 0.03:  # Orders < 3% of target
                            selected.append(item['order'])
                            totals['Qty'] += qty
                            totals['Picks'] += picks
                            totals['Hours'] += hours
                            order_difficulty = item.get('difficulty', 'Medium')
                            difficulty_counts[order_difficulty] += 1
                            remaining_orders.remove(item)
                        elif hours_improves:  # Or if it actually improves hours
                            selected.append(item['order'])
                            totals['Qty'] += qty
                            totals['Picks'] += picks
                            totals['Hours'] += hours
                            order_difficulty = item.get('difficulty', 'Medium')
                            difficulty_counts[order_difficulty] += 1
                            remaining_orders.remove(item)
                else:
                    # Over target (1-2%) - only add if it improves hours AND we need orders
                    if hours_improves and len(selected) < target_orders * 0.9:
                        selected.append(item['order'])
                        totals['Qty'] += qty
                        totals['Picks'] += picks
                        totals['Hours'] += hours
                        order_difficulty = item.get('difficulty', 'Medium')
                        difficulty_counts[order_difficulty] += 1
                        remaining_orders.remove(item)
        
        # Calculate final utilization
        utilization = {
            'Qty': totals['Qty'] / limits['Qty'] * 100 if limits['Qty'] > 0 else 0,
            'Picks': totals['Picks'] / limits['Picks'] * 100 if limits['Picks'] > 0 else 0,
            'Hours': totals['Hours'] / limits['Hours'] * 100 if limits['Hours'] > 0 else 0
        }
        
        # Calculate line distribution
        line_distribution = {
            'C1': {'count': line_counts['C1'], 'hours': line_hours['C1']},
            'C2': {'count': line_counts['C2'], 'hours': line_hours['C2']},
            'C3/4': {'count': line_counts['C3/4'], 'hours': line_hours['C3/4']},
            'Other': {'count': line_counts['Other'], 'hours': line_hours['Other']}
        }
        
        return selected, {
            'totals': totals,
            'utilization': utilization,
            'num_orders': len(selected),
            'line_distribution': line_distribution,
            'offline_count': offline_count,
            'offline_limit': limits.get('Offline Jobs', None)
        }
    
    def _calculate_balance_score(self, utilization: Dict) -> float:
        """Calculate how balanced the utilization is across all three metrics."""
        # Lower standard deviation = more balanced
        values = [utilization['Qty'], utilization['Picks'], utilization['Hours']]
        mean = sum(values) / len(values)
        variance = sum((x - mean) ** 2 for x in values) / len(values)
        std_dev = variance ** 0.5
        
        # Return negative std_dev (lower is better, so we maximize negative)
        # Also factor in total utilization
        total_util = sum(values) / len(values)
        return total_util - std_dev  # Higher is better
    
    def generate_suggestions(self, num_suggestions: int = 1, brand: str = None) -> List[Dict]:
        """Generate plan suggestions for a specific brand."""
        suggestions = []
        
        # Get limits for brand
        if brand:
            limits = self.brand_limits.get(brand, self.limits)
        else:
            limits = self.limits
            brand = 'BVI'  # Default
        
        # Generate balanced plan
        selected_orders, stats = self.optimize_plan_balanced(brand=brand, limits=limits)
        
        suggestions.append({
            'strategy': 'balanced',
            'brand': brand,
            'orders': selected_orders,
            'totals': stats['totals'],
            'utilization': stats['utilization'],
            'num_orders': stats['num_orders']
        })
        
        return suggestions
    
    def generate_all_brand_suggestions(self) -> Dict[str, List[Dict]]:
        """Generate suggestions for all brands (BVI and Malosa)."""
        all_suggestions = {}
        
        for brand in ['BVI', 'Malosa']:
            if brand in self.brand_limits:
                suggestions = self.generate_suggestions(brand=brand)
                all_suggestions[brand] = suggestions
        
        return all_suggestions
    
    def generate_multi_day_plans(self, num_days: int, brand: str = None) -> List[Dict]:
        """
        Generate multi-day plans with BALANCED distribution across days.
        
        ROUND-ROBIN APPROACH:
        1. Distribute orders round-robin (like dealing cards) for equal order counts
        2. Swap orders between days to balance hours
        3. Prioritize earlier start dates
        4. All days get similar utilization (~96-98%)
        
        Args:
            num_days: Maximum number of days to plan (will use fewer if not enough work)
            brand: Brand to plan (BVI, Malosa, etc.)
        
        Returns:
            List of day plans, each with 'day', 'orders', 'totals', 'utilization', 'num_orders'
        """
        # Get limits for brand
        if brand:
            limits = self.brand_limits.get(brand, self.limits)
            brand_orders = [o for o in self.orders if o.get('Brand', '').upper() == brand.upper()]
        else:
            limits = self.limits
            brand = 'BVI'
            brand_orders = [o for o in self.orders if o.get('Brand', '').upper() == 'BVI']
        
        if not brand_orders:
            return []
        
        hours_limit = limits['Hours']
        picks_limit = limits['Picks']
        qty_limit = limits['Qty']
        offline_limit = limits.get('Offline Jobs', float('inf'))
        
        # Prepare orders with metrics
        orders_with_metrics = []
        for order in brand_orders:
            qty = order.get('Lot Size', 0) or 0
            picks = order.get('Picks', 0) or 0
            hours = order.get('Hours', 0) or 0
            
            if qty > 0 and hours > 0:
                # Get date priority
                date_priority = 0.0
                if order.get('Start Date'):
                    try:
                        earliest_date = min(o.get('Start Date') for o in brand_orders if o.get('Start Date'))
                        if earliest_date:
                            days_diff = (order['Start Date'] - earliest_date).days
                            date_priority = 1.0 - min(days_diff / 60.0, 1.0)
                    except:
                        pass
                
                # Categorize order difficulty
                difficulty = self._categorize_order_difficulty(order, brand_orders)
                
                orders_with_metrics.append({
                    'order': order,
                    'qty': qty,
                    'picks': picks,
                    'hours': hours,
                    'date_priority': date_priority,
                    'start_date': order.get('Start Date') or datetime.max,
                    'difficulty': difficulty
                })
        
        # Calculate totals
        total_hours = sum(item['hours'] for item in orders_with_metrics)
        total_picks = sum(item['picks'] for item in orders_with_metrics)
        total_qty = sum(item['qty'] for item in orders_with_metrics)
        total_orders = len(orders_with_metrics)
        
        # Calculate actual number of days needed (based on hours limit)
        actual_days_needed = max(1, int(total_hours / hours_limit) + (1 if total_hours % hours_limit > 0 else 0))
        num_days = min(num_days, actual_days_needed)
        
        print(f"\n{'='*60}")
        print(f"MULTI-DAY PLANNING: {brand}")
        print(f"{'='*60}")
        print(f"Total orders: {total_orders}")
        print(f"Total hours: {total_hours:.1f} (need {actual_days_needed} days at {hours_limit} hrs/day)")
        print(f"Total picks: {total_picks:.0f}")
        print(f"Total qty: {total_qty:.0f}")
        print(f"Planning for {num_days} days")
        
        # Sort orders by date (earlier first)
        orders_with_metrics.sort(key=lambda x: (x['start_date'], -x['date_priority']))
        
        # Calculate targets per day for level-loading
        avg_hours_per_day = total_hours / num_days
        avg_orders_per_day = total_orders / num_days
        avg_picks_per_day = total_picks / num_days
        avg_qty_per_day = total_qty / num_days
        target_hours_per_order = avg_hours_per_day / avg_orders_per_day
        
        print(f"\nLevel-loading targets per day:")
        print(f"  Hours: {avg_hours_per_day:.1f} (limit: {hours_limit})")
        print(f"  Orders: {avg_orders_per_day:.1f}")
        print(f"  Avg hours/order: {target_hours_per_order:.1f}")
        print(f"  Picks: {avg_picks_per_day:.1f}")
        print(f"  Qty: {avg_qty_per_day:.1f}")
        
        # ========================================================================
        # PHASE 1: Initial round-robin distribution to balance order counts
        # ========================================================================
        print(f"\nPhase 1: Round-robin distribution for balanced order counts...")
        
        days = []
        for day_num in range(1, num_days + 1):
            days.append({
                'day': day_num,
                'orders': [],
                'totals': {'Qty': 0, 'Picks': 0, 'Hours': 0},
                'utilization': {'Qty': 0, 'Picks': 0, 'Hours': 0},
                'num_orders': 0,
                'line_counts': {'C1': 0, 'C2': 0, 'C3/4': 0, 'Other': 0},
                'line_hours': {'C1': 0, 'C2': 0, 'C3/4': 0, 'Other': 0},
                'offline_count': 0,
                'difficulty_counts': {'Easy': 0, 'Medium': 0, 'Difficult': 0}
            })
        
        # Distribute orders round-robin style (like dealing cards)
        # This ensures roughly equal order counts per day
        for idx, item in enumerate(orders_with_metrics):
            day_idx = idx % num_days
            self._add_order_to_day(days[day_idx], item)
        
        # Show initial distribution
        print(f"  Initial distribution (before balancing):")
        for day in days:
            print(f"    Day {day['day']}: {day['num_orders']} orders, {day['totals']['Hours']:.1f} hours ({day['totals']['Hours']/hours_limit*100:.1f}%)")
        
        # ========================================================================
        # PHASE 2: Balance hours by swapping orders between days
        # ========================================================================
        print(f"\nPhase 2: Balancing hours by swapping orders...")
        
        # Calculate current hours per day
        target_hours = min(hours_limit, avg_hours_per_day * 1.02)
        
        # Perform swaps to balance hours
        max_swap_iterations = 500
        for swap_iter in range(max_swap_iterations):
            # Find the day with most hours and the day with least hours
            days_by_hours = sorted(enumerate(days), key=lambda x: x[1]['totals']['Hours'])
            min_day_idx, min_day = days_by_hours[0]
            max_day_idx, max_day = days_by_hours[-1]
            
            hours_diff = max_day['totals']['Hours'] - min_day['totals']['Hours']
            
            # Stop if hours are balanced enough (within 5% of target)
            if hours_diff < hours_limit * 0.05:
                break
            
            # Try to find a swap that improves balance
            best_swap = None
            best_improvement = 0
            
            # Look for an order in max_day that when moved to min_day improves balance
            for i, order_in_max in enumerate(max_day['orders']):
                hours_to_move = order_in_max.get('Hours', 0) or 0
                
                new_max_hours = max_day['totals']['Hours'] - hours_to_move
                new_min_hours = min_day['totals']['Hours'] + hours_to_move
                
                # Would this improve balance?
                old_diff = abs(max_day['totals']['Hours'] - target_hours) + abs(min_day['totals']['Hours'] - target_hours)
                new_diff = abs(new_max_hours - target_hours) + abs(new_min_hours - target_hours)
                
                improvement = old_diff - new_diff
                
                if improvement > best_improvement and new_min_hours <= hours_limit * 1.05:
                    best_swap = ('move', max_day_idx, min_day_idx, i)
                    best_improvement = improvement
            
            if best_swap and best_improvement > 0.1:
                _, from_day_idx, to_day_idx, order_idx = best_swap
                
                # Perform the move
                order = days[from_day_idx]['orders'][order_idx]
                
                # Remove from source day
                days[from_day_idx]['orders'].pop(order_idx)
                days[from_day_idx]['totals']['Hours'] -= order.get('Hours', 0) or 0
                days[from_day_idx]['totals']['Qty'] -= order.get('Lot Size', 0) or 0
                days[from_day_idx]['totals']['Picks'] -= order.get('Picks', 0) or 0
                days[from_day_idx]['num_orders'] -= 1
                
                # Add to target day
                days[to_day_idx]['orders'].append(order)
                days[to_day_idx]['totals']['Hours'] += order.get('Hours', 0) or 0
                days[to_day_idx]['totals']['Qty'] += order.get('Lot Size', 0) or 0
                days[to_day_idx]['totals']['Picks'] += order.get('Picks', 0) or 0
                days[to_day_idx]['num_orders'] += 1
            else:
                # No improving swap found, try swapping two orders
                found_swap = False
                for i, order_in_max in enumerate(max_day['orders'][:20]):  # Check first 20
                    for j, order_in_min in enumerate(min_day['orders'][:20]):
                        hours_max = order_in_max.get('Hours', 0) or 0
                        hours_min = order_in_min.get('Hours', 0) or 0
                        
                        # Swap: move order_in_max to min_day, move order_in_min to max_day
                        new_max_hours = max_day['totals']['Hours'] - hours_max + hours_min
                        new_min_hours = min_day['totals']['Hours'] - hours_min + hours_max
                        
                        old_diff = abs(max_day['totals']['Hours'] - target_hours) + abs(min_day['totals']['Hours'] - target_hours)
                        new_diff = abs(new_max_hours - target_hours) + abs(new_min_hours - target_hours)
                        
                        improvement = old_diff - new_diff
                        
                        if improvement > 0.5 and new_max_hours <= hours_limit * 1.05 and new_min_hours <= hours_limit * 1.05:
                            # Perform the swap
                            # Remove both orders
                            order_max = max_day['orders'].pop(i)
                            order_min = min_day['orders'].pop(j)
                            
                            # Update totals for max_day (remove max, add min)
                            max_day['totals']['Hours'] += (order_min.get('Hours', 0) or 0) - (order_max.get('Hours', 0) or 0)
                            max_day['totals']['Qty'] += (order_min.get('Lot Size', 0) or 0) - (order_max.get('Lot Size', 0) or 0)
                            max_day['totals']['Picks'] += (order_min.get('Picks', 0) or 0) - (order_max.get('Picks', 0) or 0)
                            
                            # Update totals for min_day (remove min, add max)
                            min_day['totals']['Hours'] += (order_max.get('Hours', 0) or 0) - (order_min.get('Hours', 0) or 0)
                            min_day['totals']['Qty'] += (order_max.get('Lot Size', 0) or 0) - (order_min.get('Lot Size', 0) or 0)
                            min_day['totals']['Picks'] += (order_max.get('Picks', 0) or 0) - (order_min.get('Picks', 0) or 0)
                            
                            # Add orders to new days
                            max_day['orders'].append(order_min)
                            min_day['orders'].append(order_max)
                            
                            found_swap = True
                            break
                    if found_swap:
                        break
                
                if not found_swap:
                    break  # No more improving swaps possible
        
        print(f"  Completed {swap_iter + 1} swap iterations")
        
        # Update all day metrics after swapping
        remaining_orders = []  # No remainder needed since all orders are distributed
        
        # Finalize all days after balancing
        for day in days:
            # Recalculate line distribution
            day['line_counts'] = {'C1': 0, 'C2': 0, 'C3/4': 0, 'Other': 0}
            day['line_hours'] = {'C1': 0, 'C2': 0, 'C3/4': 0, 'Other': 0}
            day['offline_count'] = 0
            day['difficulty_counts'] = {'Easy': 0, 'Medium': 0, 'Difficult': 0}
            
            for order in day['orders']:
                order_line = self._get_line_category(order.get('Suggested Line', ''))
                day['line_counts'][order_line] += 1
                day['line_hours'][order_line] += order.get('Hours', 0) or 0
                
                order_line_raw = order.get('Suggested Line', '').strip()
                if order_line_raw.upper() == 'OFFLINE':
                    day['offline_count'] += 1
            
            day['utilization'] = {
                'Qty': day['totals']['Qty'] / qty_limit * 100 if qty_limit > 0 else 0,
                'Picks': day['totals']['Picks'] / picks_limit * 100 if picks_limit > 0 else 0,
                'Hours': day['totals']['Hours'] / hours_limit * 100 if hours_limit > 0 else 0
            }
            day['brand'] = brand
            day['day_label'] = f"Day {day['day']}"
            day['line_distribution'] = {
                'C1': {'count': day['line_counts']['C1'], 'hours': day['line_hours']['C1']},
                'C2': {'count': day['line_counts']['C2'], 'hours': day['line_hours']['C2']},
                'C3/4': {'count': day['line_counts']['C3/4'], 'hours': day['line_hours']['C3/4']},
                'Other': {'count': day['line_counts']['Other'], 'hours': day['line_hours']['Other']}
            }
            day['offline_limit'] = offline_limit
            
            print(f"\n--- Day {day['day']} ---")
            print(f"  Orders: {day['num_orders']}, Hours: {day['totals']['Hours']:.1f} ({day['utilization']['Hours']:.1f}%)")
        
        # Create Remainder with all leftover orders
        if remaining_orders:
            remainder = {
                'day': 'Remainder',
                'orders': [item['order'] for item in remaining_orders],
                'totals': {
                    'Qty': sum(item['qty'] for item in remaining_orders),
                    'Picks': sum(item['picks'] for item in remaining_orders),
                    'Hours': sum(item['hours'] for item in remaining_orders)
                },
                'utilization': {
                    'Qty': sum(item['qty'] for item in remaining_orders) / qty_limit * 100 if qty_limit > 0 else 0,
                    'Picks': sum(item['picks'] for item in remaining_orders) / picks_limit * 100 if picks_limit > 0 else 0,
                    'Hours': sum(item['hours'] for item in remaining_orders) / hours_limit * 100 if hours_limit > 0 else 0
                },
                'num_orders': len(remaining_orders),
                'line_counts': {'C1': 0, 'C2': 0, 'C3/4': 0, 'Other': 0},
                'line_hours': {'C1': 0, 'C2': 0, 'C3/4': 0, 'Other': 0},
                'offline_count': 0,
                'difficulty_counts': {'Easy': 0, 'Medium': 0, 'Difficult': 0}
            }
            
            # Calculate line distribution for remainder
            for item in remaining_orders:
                order_line = self._get_line_category(item['order'].get('Suggested Line', ''))
                remainder['line_counts'][order_line] += 1
                remainder['line_hours'][order_line] += item['hours']
                order_line_raw = item['order'].get('Suggested Line', '').strip()
                if order_line_raw.upper() == 'OFFLINE':
                    remainder['offline_count'] += 1
                order_difficulty = item.get('difficulty', 'Medium')
                remainder['difficulty_counts'][order_difficulty] += 1
            
            remainder['line_distribution'] = {
                'C1': {'count': remainder['line_counts']['C1'], 'hours': remainder['line_hours']['C1']},
                'C2': {'count': remainder['line_counts']['C2'], 'hours': remainder['line_hours']['C2']},
                'C3/4': {'count': remainder['line_counts']['C3/4'], 'hours': remainder['line_hours']['C3/4']},
                'Other': {'count': remainder['line_counts']['Other'], 'hours': remainder['line_hours']['Other']}
            }
            remainder['offline_limit'] = offline_limit
            remainder['brand'] = brand
            remainder['day_label'] = 'Remainder'
            
            print(f"\n--- Remainder ---")
            print(f"  {remainder['num_orders']} orders, {remainder['totals']['Hours']:.1f} hours ({remainder['utilization']['Hours']:.1f}%)")
            
            days.append(remainder)
        
        return days
    
    def _add_order_to_day(self, day: Dict, item: Dict):
        """Helper to add an order to a day and update all tracking."""
        day['orders'].append(item['order'])
        day['totals']['Qty'] += item['qty']
        day['totals']['Picks'] += item['picks']
        day['totals']['Hours'] += item['hours']
        day['num_orders'] += 1
        
        order_line = self._get_line_category(item['order'].get('Suggested Line', ''))
        day['line_counts'][order_line] += 1
        day['line_hours'][order_line] += item['hours']
        
        order_difficulty = item.get('difficulty', 'Medium')
        day['difficulty_counts'][order_difficulty] += 1
        
        order_line_raw = item['order'].get('Suggested Line', '').strip()
        if order_line_raw.upper() == 'OFFLINE':
            day['offline_count'] += 1
    
    def generate_multi_day_plans_with_scenarios(self, num_days: int, brand: str = None) -> Dict:
        """
        Generate multiple scenarios for multi-day planning and compare them.
        
        Scenarios tested:
        1. Date Priority First - Prioritize earlier start dates above all
        2. Balanced - Balance all metrics equally
        3. Hours Optimized - Maximize hours utilization per day
        
        Returns dict with 'best' scenario and 'all_scenarios' for comparison.
        """
        scenarios = []
        
        # Scenario 1: Date Priority First
        print("\n" + "="*60)
        print("SCENARIO 1: DATE PRIORITY FIRST")
        print("="*60)
        scenario1 = self._run_scenario(num_days, brand, strategy='date_first')
        scenario1['name'] = 'Date Priority First'
        scenario1['description'] = 'Strongly prioritizes earlier start dates'
        scenarios.append(scenario1)
        
        # Scenario 2: Balanced
        print("\n" + "="*60)
        print("SCENARIO 2: BALANCED")
        print("="*60)
        scenario2 = self._run_scenario(num_days, brand, strategy='balanced')
        scenario2['name'] = 'Balanced'
        scenario2['description'] = 'Balances hours, dates, lines, and difficulty'
        scenarios.append(scenario2)
        
        # Scenario 3: Hours Optimized (pack each day as efficiently as possible)
        print("\n" + "="*60)
        print("SCENARIO 3: HOURS OPTIMIZED")
        print("="*60)
        scenario3 = self._run_scenario(num_days, brand, strategy='hours_first')
        scenario3['name'] = 'Hours Optimized'
        scenario3['description'] = 'Maximizes hours utilization per day'
        scenarios.append(scenario3)
        
        # Score each scenario
        for scenario in scenarios:
            scenario['score'] = self._score_scenario(scenario)
        
        # Find best scenario
        best_scenario = max(scenarios, key=lambda s: s['score'])
        
        # Print comparison
        print("\n" + "="*60)
        print("SCENARIO COMPARISON")
        print("="*60)
        for scenario in scenarios:
            complete_days = [d for d in scenario['days'] if d.get('day') != 'Remainder']
            remainder = [d for d in scenario['days'] if d.get('day') == 'Remainder']
            
            if complete_days:
                avg_hours_util = sum(d['utilization']['Hours'] for d in complete_days) / len(complete_days)
                min_hours_util = min(d['utilization']['Hours'] for d in complete_days)
                max_hours_util = max(d['utilization']['Hours'] for d in complete_days)
                hours_std = self._calculate_std([d['utilization']['Hours'] for d in complete_days])
            else:
                avg_hours_util = min_hours_util = max_hours_util = hours_std = 0
            
            remainder_hours = remainder[0]['totals']['Hours'] if remainder else 0
            
            print(f"\n{scenario['name']} (Score: {scenario['score']:.1f})")
            print(f"  Complete days: {len(complete_days)}")
            print(f"  Hours util: avg={avg_hours_util:.1f}%, min={min_hours_util:.1f}%, max={max_hours_util:.1f}%, std={hours_std:.2f}")
            print(f"  Remainder hours: {remainder_hours:.1f}")
            if scenario == best_scenario:
                print(f"  *** BEST SCENARIO ***")
        
        return {
            'best': best_scenario,
            'all_scenarios': scenarios
        }
    
    def _run_scenario(self, num_days: int, brand: str, strategy: str) -> Dict:
        """Run a specific planning scenario."""
        # Get limits for brand
        if brand:
            limits = self.brand_limits.get(brand, self.limits)
            brand_orders = [o for o in self.orders if o.get('Brand', '').upper() == brand.upper()]
        else:
            limits = self.limits
            brand = 'BVI'
            brand_orders = [o for o in self.orders if o.get('Brand', '').upper() == 'BVI']
        
        if not brand_orders:
            return {'days': [], 'strategy': strategy}
        
        hours_limit = limits['Hours']
        picks_limit = limits['Picks']
        qty_limit = limits['Qty']
        offline_limit = limits.get('Offline Jobs', float('inf'))
        
        # Prepare orders with metrics
        orders_with_metrics = []
        for order in brand_orders:
            qty = order.get('Lot Size', 0) or 0
            picks = order.get('Picks', 0) or 0
            hours = order.get('Hours', 0) or 0
            
            if qty > 0 and hours > 0:
                date_priority = 0.0
                if order.get('Start Date'):
                    try:
                        earliest_date = min(o.get('Start Date') for o in brand_orders if o.get('Start Date'))
                        if earliest_date:
                            days_diff = (order['Start Date'] - earliest_date).days
                            date_priority = 1.0 - min(days_diff / 60.0, 1.0)
                    except:
                        pass
                
                difficulty = self._categorize_order_difficulty(order, brand_orders)
                
                orders_with_metrics.append({
                    'order': order,
                    'qty': qty,
                    'picks': picks,
                    'hours': hours,
                    'date_priority': date_priority,
                    'start_date': order.get('Start Date') or datetime.max,
                    'difficulty': difficulty
                })
        
        total_hours = sum(item['hours'] for item in orders_with_metrics)
        actual_days_needed = max(1, int(total_hours / hours_limit) + (1 if total_hours % hours_limit > 0 else 0))
        num_days = min(num_days, actual_days_needed)
        
        # Sort based on strategy
        if strategy == 'date_first':
            orders_with_metrics.sort(key=lambda x: (x['start_date'], -x['date_priority']))
        elif strategy == 'hours_first':
            orders_with_metrics.sort(key=lambda x: (-x['hours'], x['start_date']))
        else:  # balanced
            orders_with_metrics.sort(key=lambda x: (x['start_date'], -x['hours'] * 0.3 - x['date_priority'] * 0.7))
        
        remaining_orders = orders_with_metrics.copy()
        days = []
        
        for day_num in range(1, num_days + 1):
            if not remaining_orders:
                break
            
            day = {
                'day': day_num,
                'orders': [],
                'totals': {'Qty': 0, 'Picks': 0, 'Hours': 0},
                'utilization': {'Qty': 0, 'Picks': 0, 'Hours': 0},
                'num_orders': 0,
                'line_counts': {'C1': 0, 'C2': 0, 'C3/4': 0, 'Other': 0},
                'line_hours': {'C1': 0, 'C2': 0, 'C3/4': 0, 'Other': 0},
                'offline_count': 0,
                'difficulty_counts': {'Easy': 0, 'Medium': 0, 'Difficult': 0}
            }
            
            remaining_days = num_days - day_num + 1
            remaining_hours = sum(item['hours'] for item in remaining_orders)
            target_hours = min(hours_limit, remaining_hours / remaining_days * 1.02)
            target_hours = max(target_hours, hours_limit * 0.95)
            
            # Fill day based on strategy
            max_iterations = len(remaining_orders) * 2
            iteration = 0
            
            while remaining_orders and iteration < max_iterations:
                iteration += 1
                
                if day['totals']['Hours'] >= target_hours * 0.98:
                    if day['totals']['Hours'] >= hours_limit * 0.995:
                        break
                
                best_order = None
                best_score = -float('inf')
                
                for item in remaining_orders:
                    hours = item['hours']
                    new_hours = day['totals']['Hours'] + hours
                    
                    if new_hours > hours_limit * 1.05:
                        continue
                    
                    order_line_raw = item['order'].get('Suggested Line', '').strip()
                    is_offline = order_line_raw.upper() == 'OFFLINE'
                    if is_offline and day['offline_count'] >= offline_limit:
                        continue
                    
                    # Score based on strategy
                    if strategy == 'date_first':
                        score = item['date_priority'] * 100
                        hours_distance = abs(target_hours - new_hours)
                        if new_hours <= target_hours:
                            score += (target_hours - hours_distance) * 0.5
                        else:
                            score -= hours_distance * 2
                    elif strategy == 'hours_first':
                        hours_distance_before = abs(target_hours - day['totals']['Hours'])
                        hours_distance_after = abs(target_hours - new_hours)
                        score = (hours_distance_before - hours_distance_after) * 20
                        if new_hours <= target_hours:
                            score += 10
                        score += item['date_priority'] * 10
                    else:  # balanced
                        hours_distance_before = abs(target_hours - day['totals']['Hours'])
                        hours_distance_after = abs(target_hours - new_hours)
                        hours_score = (hours_distance_before - hours_distance_after) * 10
                        if new_hours <= target_hours:
                            hours_score += 5
                        date_score = item['date_priority'] * 30
                        
                        order_line = self._get_line_category(item['order'].get('Suggested Line', ''))
                        line_score = 0
                        if order_line in ['C1', 'C2', 'C3/4']:
                            total_line = sum(day['line_counts'][l] for l in ['C1', 'C2', 'C3/4'])
                            if total_line > 0:
                                ratio = day['line_counts'][order_line] / total_line
                                if ratio < 0.25:
                                    line_score = 15
                        
                        score = hours_score + date_score + line_score
                    
                    if score > best_score:
                        best_score = score
                        best_order = item
                
                if best_order:
                    self._add_order_to_day(day, best_order)
                    remaining_orders.remove(best_order)
                else:
                    for item in remaining_orders:
                        if day['totals']['Hours'] + item['hours'] <= hours_limit * 1.05:
                            order_line_raw = item['order'].get('Suggested Line', '').strip()
                            is_offline = order_line_raw.upper() == 'OFFLINE'
                            if not (is_offline and day['offline_count'] >= offline_limit):
                                self._add_order_to_day(day, item)
                                remaining_orders.remove(item)
                                break
                    else:
                        break
            
            # Try to reach exactly 100%
            if day['totals']['Hours'] < hours_limit * 0.98:
                for item in sorted(remaining_orders, key=lambda x: x['hours']):
                    if day['totals']['Hours'] + item['hours'] <= hours_limit * 1.02:
                        order_line_raw = item['order'].get('Suggested Line', '').strip()
                        is_offline = order_line_raw.upper() == 'OFFLINE'
                        if not (is_offline and day['offline_count'] >= offline_limit):
                            self._add_order_to_day(day, item)
                            remaining_orders.remove(item)
                            if day['totals']['Hours'] >= hours_limit * 0.98:
                                break
            
            day['utilization'] = {
                'Qty': day['totals']['Qty'] / qty_limit * 100 if qty_limit > 0 else 0,
                'Picks': day['totals']['Picks'] / picks_limit * 100 if picks_limit > 0 else 0,
                'Hours': day['totals']['Hours'] / hours_limit * 100 if hours_limit > 0 else 0
            }
            day['brand'] = brand
            day['day_label'] = f"Day {day_num}"
            day['line_distribution'] = {
                'C1': {'count': day['line_counts']['C1'], 'hours': day['line_hours']['C1']},
                'C2': {'count': day['line_counts']['C2'], 'hours': day['line_hours']['C2']},
                'C3/4': {'count': day['line_counts']['C3/4'], 'hours': day['line_hours']['C3/4']},
                'Other': {'count': day['line_counts']['Other'], 'hours': day['line_hours']['Other']}
            }
            day['offline_limit'] = offline_limit
            
            if day['num_orders'] > 0:
                days.append(day)
        
        # Remainder
        if remaining_orders:
            remainder = {
                'day': 'Remainder',
                'orders': [item['order'] for item in remaining_orders],
                'totals': {
                    'Qty': sum(item['qty'] for item in remaining_orders),
                    'Picks': sum(item['picks'] for item in remaining_orders),
                    'Hours': sum(item['hours'] for item in remaining_orders)
                },
                'utilization': {
                    'Qty': sum(item['qty'] for item in remaining_orders) / qty_limit * 100 if qty_limit > 0 else 0,
                    'Picks': sum(item['picks'] for item in remaining_orders) / picks_limit * 100 if picks_limit > 0 else 0,
                    'Hours': sum(item['hours'] for item in remaining_orders) / hours_limit * 100 if hours_limit > 0 else 0
                },
                'num_orders': len(remaining_orders),
                'line_counts': {'C1': 0, 'C2': 0, 'C3/4': 0, 'Other': 0},
                'line_hours': {'C1': 0, 'C2': 0, 'C3/4': 0, 'Other': 0},
                'offline_count': 0,
                'difficulty_counts': {'Easy': 0, 'Medium': 0, 'Difficult': 0}
            }
            
            for item in remaining_orders:
                order_line = self._get_line_category(item['order'].get('Suggested Line', ''))
                remainder['line_counts'][order_line] += 1
                remainder['line_hours'][order_line] += item['hours']
            
            remainder['line_distribution'] = {
                'C1': {'count': remainder['line_counts']['C1'], 'hours': remainder['line_hours']['C1']},
                'C2': {'count': remainder['line_counts']['C2'], 'hours': remainder['line_hours']['C2']},
                'C3/4': {'count': remainder['line_counts']['C3/4'], 'hours': remainder['line_hours']['C3/4']},
                'Other': {'count': remainder['line_counts']['Other'], 'hours': remainder['line_hours']['Other']}
            }
            remainder['offline_limit'] = offline_limit
            remainder['brand'] = brand
            remainder['day_label'] = 'Remainder'
            days.append(remainder)
        
        return {'days': days, 'strategy': strategy}
    
    def _score_scenario(self, scenario: Dict) -> float:
        """
        Score a scenario based on:
        1. Hours utilization (MOST IMPORTANT - 100% per day is the goal)
        2. Balance across days (low standard deviation)
        3. Minimal remainder
        4. Date priority (earlier dates scheduled first)
        """
        days = scenario.get('days', [])
        complete_days = [d for d in days if d.get('day') != 'Remainder']
        remainder = [d for d in days if d.get('day') == 'Remainder']
        
        if not complete_days:
            return 0
        
        # 1. Hours utilization score (target: 100% each day)
        hours_utils = [d['utilization']['Hours'] for d in complete_days]
        
        # Penalty for each day not at 100%
        hours_penalty = sum(abs(100 - h) for h in hours_utils)
        hours_score = 100 - hours_penalty * 2  # Strong penalty
        
        # 2. Balance score (low std dev across days)
        if len(hours_utils) > 1:
            hours_std = self._calculate_std(hours_utils)
            balance_score = 50 - hours_std * 5
        else:
            balance_score = 50
        
        # 3. Remainder penalty
        remainder_hours = remainder[0]['totals']['Hours'] if remainder else 0
        remainder_penalty = min(remainder_hours / 100, 50)  # Cap penalty
        
        # 4. Date priority bonus (check if earlier dates are in earlier days)
        date_score = 0
        for i, day in enumerate(complete_days):
            avg_date_priority = sum(1 for o in day['orders'] if o.get('Start Date')) / max(day['num_orders'], 1)
            date_score += avg_date_priority * (len(complete_days) - i)  # Earlier days weighted more
        
        total_score = hours_score + balance_score - remainder_penalty + date_score
        return max(0, total_score)
    
    def _calculate_std(self, values: list) -> float:
        """Calculate standard deviation."""
        if len(values) < 2:
            return 0
        mean = sum(values) / len(values)
        variance = sum((x - mean) ** 2 for x in values) / len(values)
        return variance ** 0.5
    
    def export_to_excel(self, suggestions: List[Dict], output_path: str = 'daily_plan_suggestions.xlsx'):
        """Export suggestions to Excel file."""
        wb = openpyxl.Workbook()
        
        # Check if this is multi-day (all suggestions have 'day' key)
        is_multi_day = len(suggestions) > 1 and 'day' in suggestions[0]
        
        if is_multi_day:
            # Combine all days into a single sheet with Day column
            ws = wb.active
            ws.title = "Multi-Day Plan"
            
            # Write summary for each day
            row = 1
            ws['A1'] = 'Day Summary'
            row += 1
            
            for suggestion in suggestions:
                day_label = suggestion.get('day_label', f"Day {suggestion.get('day', '?')}")
                ws[f'A{row}'] = day_label
                ws[f'B{row}'] = 'Orders'
                ws[f'C{row}'] = suggestion['num_orders']
                ws[f'D{row}'] = 'Qty'
                ws[f'E{row}'] = suggestion['totals']['Qty']
                ws[f'F{row}'] = 'Picks'
                ws[f'G{row}'] = suggestion['totals']['Picks']
                ws[f'H{row}'] = 'Hours'
                ws[f'I{row}'] = suggestion['totals']['Hours']
                ws[f'J{row}'] = 'Hours %'
                ws[f'K{row}'] = f"{suggestion['utilization']['Hours']:.1f}%"
                row += 1
            
            row += 1
            ws[f'A{row}'] = 'Orders:'
            row += 1
            
            # Collect all orders with day labels
            all_orders = []
            for suggestion in suggestions:
                day_label = suggestion.get('day_label', f"Day {suggestion.get('day', '?')}")
                for order in suggestion['orders']:
                    order_with_day = order.copy()
                    order_with_day['Day'] = day_label
                    all_orders.append(order_with_day)
            
            if all_orders:
                # Write headers with Day column first
                headers = ['Day'] + [k for k in all_orders[0].keys() if k != 'Day']
                for col_idx, header in enumerate(headers, 1):
                    ws.cell(row=row, column=col_idx, value=header)
                row += 1
                
                # Write orders
                for order in all_orders:
                    col_idx = 1
                    # Write Day first
                    ws.cell(row=row, column=col_idx, value=order.get('Day', ''))
                    col_idx += 1
                    # Write other fields
                    for header in headers[1:]:
                        value = order.get(header, '')
                        if isinstance(value, datetime):
                            value = value.strftime('%Y-%m-%d')
                        ws.cell(row=row, column=col_idx, value=value)
                        col_idx += 1
                    row += 1
        else:
            # Original single-day behavior
            for suggestion in suggestions:
                # Determine sheet name - use day label if available, otherwise strategy
                if 'day_label' in suggestion:
                    sheet_name = suggestion['day_label']
                elif 'day' in suggestion:
                    sheet_name = f"Day {suggestion['day']}"
                else:
                    sheet_name = suggestion.get('strategy', 'Plan').title()
                
                # Limit sheet name to 31 characters (Excel limit)
                if len(sheet_name) > 31:
                    sheet_name = sheet_name[:31]
                
                # Create sheet for this suggestion
                ws = wb.create_sheet(title=sheet_name)
                
                # Write summary
                if 'day_label' in suggestion:
                    ws['A1'] = 'Day'
                    ws['B1'] = suggestion['day_label']
                else:
                    ws['A1'] = 'Strategy'
                    ws['B1'] = suggestion.get('strategy', 'N/A')
                
                if 'brand' in suggestion:
                    ws['A2'] = 'Brand'
                    ws['B2'] = suggestion['brand']
                    row_offset = 1
                else:
                    row_offset = 0
                
                ws[f'A{3+row_offset}'] = 'Number of Orders'
                ws[f'B{3+row_offset}'] = suggestion['num_orders']
                ws[f'A{4+row_offset}'] = 'Total Qty'
                ws[f'B{4+row_offset}'] = suggestion['totals']['Qty']
                ws[f'A{5+row_offset}'] = 'Total Picks'
                ws[f'B{5+row_offset}'] = suggestion['totals']['Picks']
                ws[f'A{6+row_offset}'] = 'Total Hours'
                ws[f'B{6+row_offset}'] = suggestion['totals']['Hours']
                ws[f'A{7+row_offset}'] = 'Qty Utilization %'
                ws[f'B{7+row_offset}'] = f"{suggestion['utilization']['Qty']:.1f}%"
                ws[f'A{8+row_offset}'] = 'Picks Utilization %'
                ws[f'B{8+row_offset}'] = f"{suggestion['utilization']['Picks']:.1f}%"
                ws[f'A{9+row_offset}'] = 'Hours Utilization %'
                ws[f'B{9+row_offset}'] = f"{suggestion['utilization']['Hours']:.1f}%"
                
                # Write orders
                if suggestion['orders']:
                    headers = list(suggestion['orders'][0].keys())
                    ws[f'A{11+row_offset}'] = 'Orders:'
                    for col_idx, header in enumerate(headers, 1):
                        ws.cell(row=12+row_offset, column=col_idx, value=header)
                    
                    for row_idx, order in enumerate(suggestion['orders'], 13+row_offset):
                        for col_idx, header in enumerate(headers, 1):
                            value = order.get(header, '')
                            if isinstance(value, datetime):
                                value = value.strftime('%Y-%m-%d')
                            ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Remove default sheet if we created new ones
        if not is_multi_day and 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        wb.save(output_path)
        print(f"Exported to {output_path}")
    
    def export_to_csv(self, suggestions: List[Dict], output_path: str = 'daily_plan_suggestions.csv'):
        """Export suggestions to CSV file."""
        if not suggestions:
            return
        
        # Check if this is multi-day (all suggestions have 'day' key)
        is_multi_day = len(suggestions) > 1 and 'day' in suggestions[0]
        
        if is_multi_day:
            # Combine all days into a single CSV with Day column
            all_orders = []
            for suggestion in suggestions:
                day_label = suggestion.get('day_label', f"Day {suggestion.get('day', '?')}")
                for order in suggestion['orders']:
                    order_with_day = order.copy()
                    order_with_day['Day'] = day_label
                    all_orders.append(order_with_day)
            
            if all_orders:
                # Get fieldnames with Day first
                fieldnames = ['Day'] + [k for k in all_orders[0].keys() if k != 'Day']
                
                with open(output_path, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.DictWriter(f, fieldnames=fieldnames)
                    writer.writeheader()
                    
                    for order in all_orders:
                        row = {}
                        for key in fieldnames:
                            value = order.get(key, '')
                            if isinstance(value, datetime):
                                row[key] = value.strftime('%Y-%m-%d')
                            else:
                                row[key] = value
                        writer.writerow(row)
                print(f"Exported to {output_path}")
        else:
            # Single suggestion - export normally
            suggestion = suggestions[0]
            
            with open(output_path, 'w', newline='', encoding='utf-8') as f:
                if suggestion['orders']:
                    fieldnames = list(suggestion['orders'][0].keys())
                    writer = csv.DictWriter(f, fieldnames=fieldnames)
                    writer.writeheader()
                    
                    for order in suggestion['orders']:
                        # Convert datetime to string
                        row = {}
                        for key, value in order.items():
                            if isinstance(value, datetime):
                                row[key] = value.strftime('%Y-%m-%d')
                            else:
                                row[key] = value
                        writer.writerow(row)
            print(f"Exported to {output_path}")


def main():
    """Main function to run the optimizer."""
    import sys
    
    template_path = "Daily Planning Template.xlsm"
    
    # Load data directly from Excel template
    optimizer = DailyPlanOptimizer(template_path=template_path)
    optimizer.load_data()
    
    print("\n" + "="*60)
    print("MULTI-DAY PLANNING OPTIMIZER")
    print("="*60)
    print("\nThis optimizer will:")
    print("  1. GUARANTEE 100% hours utilization per day")
    print("  2. Balance order counts across days")
    print("  3. Level-load hours, picks, qty across days")
    print("  4. Prioritize earlier start dates")
    print("  5. Balance line distribution (C1:C2:C3/4)")
    
    for brand in ['BVI', 'Malosa']:
        if brand in optimizer.brand_limits:
            # Calculate estimated max days
            limits = optimizer.brand_limits[brand]
            brand_orders = [o for o in optimizer.orders if o.get('Brand', '').upper() == brand.upper()]
            
            if not brand_orders:
                print(f"\nNo orders found for {brand}")
                continue
            
            total_hours = sum(o.get('Hours', 0) or 0 for o in brand_orders)
            hours_per_day = limits['Hours']
            estimated_max_days = max(1, int(total_hours / hours_per_day) + 2)
            
            # Run the optimized multi-day planning
            day_plans = optimizer.generate_multi_day_plans(estimated_max_days, brand=brand)
            
            if day_plans:
                complete_days = [d for d in day_plans if d.get('day') != 'Remainder']
                num_complete_days = len(complete_days)
                
                print(f"\n{'='*60}")
                print(f"RESULTS: {brand}")
                print(f"{'='*60}")
                print(f"\nGenerated {num_complete_days} complete day(s)")
                
                # Summary table
                print(f"\n{'Day':<12} {'Orders':<8} {'Qty':<10} {'Picks':<10} {'Hours':<10} {'Hours %':<10}")
                print("-" * 70)
                
                for day_plan in day_plans:
                    day_label = day_plan.get('day_label', f"Day {day_plan.get('day', '?')}")
                    print(f"{day_label:<12} {day_plan['num_orders']:<8} "
                          f"{day_plan['totals']['Qty']:<10.0f} "
                          f"{day_plan['totals']['Picks']:<10.0f} "
                          f"{day_plan['totals']['Hours']:<10.1f} "
                          f"{day_plan['utilization']['Hours']:<10.1f}%")
                
                # Calculate totals
                total_orders = sum(d['num_orders'] for d in day_plans)
                total_qty = sum(d['totals']['Qty'] for d in day_plans)
                total_picks = sum(d['totals']['Picks'] for d in day_plans)
                total_hours_planned = sum(d['totals']['Hours'] for d in day_plans)
                
                print("-" * 70)
                print(f"{'TOTAL':<12} {total_orders:<8} "
                      f"{total_qty:<10.0f} "
                      f"{total_picks:<10.0f} "
                      f"{total_hours_planned:<10.1f}")
                
                # Balance metrics
                if len(complete_days) > 1:
                    hours_utils = [d['utilization']['Hours'] for d in complete_days]
                    orders_counts = [d['num_orders'] for d in complete_days]
                    
                    avg_hours = sum(hours_utils) / len(hours_utils)
                    min_hours = min(hours_utils)
                    max_hours = max(hours_utils)
                    std_hours = optimizer._calculate_std(hours_utils)
                    
                    avg_orders = sum(orders_counts) / len(orders_counts)
                    min_orders = min(orders_counts)
                    max_orders = max(orders_counts)
                    std_orders = optimizer._calculate_std(orders_counts)
                    
                    print(f"\nBalance Metrics (Complete Days):")
                    print(f"  Hours: avg={avg_hours:.1f}%, min={min_hours:.1f}%, max={max_hours:.1f}%, std={std_hours:.2f}")
                    print(f"  Orders: avg={avg_orders:.1f}, min={min_orders}, max={max_orders}, std={std_orders:.2f}")
                
                # Line distribution summary
                print(f"\nLine Distribution per Day:")
                for day_plan in complete_days:
                    day_label = day_plan.get('day_label', f"Day {day_plan.get('day', '?')}")
                    line_dist = day_plan.get('line_distribution', {})
                    c1 = line_dist.get('C1', {}).get('count', 0)
                    c2 = line_dist.get('C2', {}).get('count', 0)
                    c34 = line_dist.get('C3/4', {}).get('count', 0)
                    other = line_dist.get('Other', {}).get('count', 0)
                    print(f"  {day_label}: C1={c1}, C2={c2}, C3/4={c34}, Other={other}")
                
                # Create output directory
                output_dir = 'output'
                if not os.path.exists(output_dir):
                    os.makedirs(output_dir)
                
                # Generate timestamp
                timestamp = datetime.now().strftime('%Y%m%d%H%M')
                
                # Export
                brand_lower = brand.lower()
                excel_filename = f'{timestamp}-{brand_lower}-plan-suggestion.xlsx'
                excel_path = os.path.join(output_dir, excel_filename)
                
                optimizer.export_to_excel(day_plans, excel_path)
                print(f"\nExported to: {excel_path}")
            else:
                print(f"No plans generated for {brand}")
    
    print("\n" + "="*60)
    print("Done! Check the generated Excel files for each brand.")
    print("="*60)


if __name__ == "__main__":
    main()